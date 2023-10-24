from django.shortcuts import render
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.urls import reverse
from clientsMetalprotec.models import clientSystem
from productsMetalprotec.models import productSystem, storeSystem, storexproductSystem
from django.contrib.auth.models import User
from servicesMetalprotec.models import serviceSystem
from usersMetalprotec.models import extendedUser
from settingsMetalprotec.models import endpointSystem
from .models import quotationSystem, quotationClientData, quotationProductData, quotationSellerData, quotationServiceData, guideSystem, creditNoteSystem, invoiceSystem, billSystem
from dateutil.parser import parse
import json
import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from decimal import Decimal, DecimalException,getcontext
import environ
import os
import requests
from base64 import b64decode
from bs4 import BeautifulSoup
from stockManagment.models import outcomingItemsRegisterInfo
from django.db.models import Q
import pandas as pd
import openpyxl
import time

env = environ.Env()
env.read_env()

token_metalprotec='HUR89LVdEfuKRdtpIqHYEbj5+3YFgJxBi2ecFzzQfVB5AAERhObWzBNga6NjSgH7'

getcontext().prec = 10

# Create your views here.
def quotationsMetalprotec(request):
    return render(request,'quotationsMetalprotec.html',{
        'quotationsSystem':quotationSystem.objects.filter(endpointQuotation=request.user.extendeduser.endpointUser).order_by('-codeQuotation')
    })

def guidesMetalprotec(request):
    return render(request,'guidesMetalprotec.html',{
        'guidesSystem':guideSystem.objects.filter(endpointGuide=request.user.extendeduser.endpointUser).order_by('-codeGuide')
    })

def billsMetalprotec(request):
    return render(request,'billsMetalprotec.html',{
        'billsSystem':billSystem.objects.filter(endpointBill=request.user.extendeduser.endpointUser).order_by('-codeBill')
    })

def invoicesMetalprotec(request):
    return render(request,'invoicesMetalprotec.html',{
        'invoicesSystem':invoiceSystem.objects.filter(endpointInvoice=request.user.extendeduser.endpointUser).order_by('-codeInvoice')
    })

def creditNotesMetalprotec(request):
    return render(request,'creditNotesMetalprotec.html',{
        'creditNotesSystem':creditNoteSystem.objects.filter(endpointCreditNote=request.user.extendeduser.endpointUser).order_by('-codeCreditNote')
    })

def editDataQuotation(request,idQuotation):
    quotationObject = quotationSystem.objects.get(id=idQuotation)
    return render(request,'editQuotation.html',{
        'clientsSystem':clientSystem.objects.filter(endpointClient=request.user.extendeduser.endpointUser).order_by('id'),
        'usersSystem':extendedUser.objects.filter(endpointUser=request.user.extendeduser.endpointUser).order_by('id'),
        'productsSystem':productSystem.objects.filter(endpointProduct=request.user.extendeduser.endpointUser).order_by('id'),
        'servicesSystem':serviceSystem.objects.filter(endpointService=request.user.extendeduser.endpointUser).order_by('id'),
        'editQuotation':quotationObject,
    })


def editDataGuide(request,idGuide):
    guideObject = guideSystem.objects.get(id=idGuide)
    return render(request,'editGuide.html',{
        'editGuide':guideObject,
    })

def editDataBill(request,idBill):
    billObject = billSystem.objects.get(id=idBill)
    return render(request,'editBill.html',{
        'editBill':billObject,
    })

def editDataInvoice(request,idInvoice):
    invoiceObject = invoiceSystem.objects.get(id=idInvoice)
    return render(request,'editInvoice.html',{
        'editInvoice':invoiceObject,
    })

def newQuotation(request):
    exchangeRate = getExchangeRate()
    if request.method == 'POST':
        quotationInfo = json.load(request)
        productsData = quotationInfo.get('productsData')
        servicesData = quotationInfo.get('servicesData')
        clientData = quotationInfo.get('clientData')
        sellerData = quotationInfo.get('sellerData')
        quotationData = quotationInfo.get('quotationData')
        documentOptions = quotationInfo.get('documentOptions')
        newQuotationItem = quotationSystem.objects.create(
            endpointQuotation=request.user.extendeduser.endpointUser,
            showDiscount = documentOptions.get('showDiscount'),
            showUnitPrice = documentOptions.get('showUnitPrice'),
            showSellPrice = documentOptions.get('showSellPrice'),
            relatedDocumentQuotation = quotationData.get('relatedDocumentQuotation'),
            commentQuotation = quotationData.get('commentQuotation'),
            quotesQuotation = quotationData.get('quotesQuotation'),
            expirationCredit = quotationData.get('expirationCredit'),
            expirationQuotation = quotationData.get('expirationQuotation'),
            erBuy = quotationData.get('erBuy'),
            erSel = quotationData.get('erSel'),
            stateQuotation = 'GENERADA',
            paymentQuotation = quotationData.get('paymentQuotation'),
            dateQuotation = parse(quotationData.get('dateQuotation')),
            typeQuotation = quotationData.get('typeItems'),
            currencyQuotation = quotationData.get('currencyQuotation'),
        )

        endpointData = endpointSystem.objects.get(id=request.user.extendeduser.endpointUser.id)
        numberQuotation = endpointData.nroCoti
        serieQuotation = endpointData.serieCoti
        newQuotationItem.numberQuotation = numberQuotation
        endpointData.nroCoti = str(int(numberQuotation) + 1)
        endpointData.save()
        newQuotationItem.save()
        codeQuotation = numberQuotation
        while len(codeQuotation) < 4:
            codeQuotation = '0' + codeQuotation
        codeQuotation = f'{serieQuotation}-{codeQuotation}'
        newQuotationItem.codeQuotation = codeQuotation
        newQuotationItem.save()

        clientQuotation = clientSystem.objects.get(id=clientData.get('idClient'))
        idClient=clientData.get('idClient')
        identificationClient=clientData.get('identificationClient')
        documentClient=clientData.get('documentClient')
        typeClient=clientData.get('typeClient')
        emailClient=clientData.get('emailClient')
        contactClient=clientData.get('contactClient')
        phoneClient=clientData.get('phoneClient')
        legalAddressClient=clientData.get('legalAddressClient')
        deliveryAddressClient=clientData.get('deliveryAddress')
        print(deliveryAddressClient)
        dataClientQuotation = [
            idClient,
            identificationClient,
            documentClient,
            typeClient,
            emailClient,
            contactClient,
            phoneClient,
            legalAddressClient,
            deliveryAddressClient,
        ]
        print(dataClientQuotation)
        quotationClientInfo = quotationClientData.objects.create(
            asociatedQuotation=newQuotationItem,
            asociatedClient=clientQuotation,
            dataClientQuotation=dataClientQuotation,
        )
        quotationClientInfo.save()

        sellerQuotation = extendedUser.objects.get(id=sellerData.get('idSeller')).asociatedUser
        idSeller=sellerData.get('idSeller')
        nameSeller=sellerData.get('nameSeller')
        codeSeller=sellerData.get('codeSeller')
        phoneSeller=sellerData.get('phoneSeller')
        dataUserQuotation = [
            idSeller,
            nameSeller,
            codeSeller,
            phoneSeller,
        ]
        quotationSellerInfo = quotationSellerData.objects.create(
            asociatedQuotation=newQuotationItem,
            asociatedUser=sellerQuotation,
            dataUserQuotation=dataUserQuotation,
        )
        quotationSellerInfo.save()

        for productInfo in productsData:
            asociatedProduct = productSystem.objects.get(id=productInfo[0])
            productInfo.append(asociatedProduct.weightProduct)
            quotationProductData.objects.create(
                asociatedQuotation=newQuotationItem,
                asociatedProduct=asociatedProduct,
                dataProductQuotation=productInfo,
            )

        for serviceInfo in servicesData:
            quotationServiceData.objects.create(
                asociatedQuotation=newQuotationItem,
                asociatedService=serviceSystem.objects.get(id=serviceInfo[0]),
                dataServiceQuotation=serviceInfo,
            )
        time.sleep(0.5)
        return JsonResponse({
            'metalprotec':'200',
        })
    return render(request,'newQuotation.html',{
        'clientsSystem':clientSystem.objects.filter(endpointClient=request.user.extendeduser.endpointUser).order_by('id'),
        'usersSystem':extendedUser.objects.filter(endpointUser=request.user.extendeduser.endpointUser).order_by('id'),
        'productsSystem':productSystem.objects.filter(endpointProduct=request.user.extendeduser.endpointUser).order_by('id'),
        'servicesSystem':serviceSystem.objects.filter(endpointService=request.user.extendeduser.endpointUser).order_by('id'),
        'exchangeRate':exchangeRate
    })


def createGuideFromQuotation(request, idQuotation):
    quotationObject = quotationSystem.objects.get(id=idQuotation)
    commentGuide = quotationObject.commentQuotation
    quotationObject.stateQuotation = 'EMITIDA'
    quotationObject.save()
    dateGuide = datetime.datetime.today()
    dateGivenGoods = datetime.datetime.today()
    endpointData = endpointSystem.objects.get(id=request.user.extendeduser.endpointUser.id)
    nroGuide = endpointData.nroGuia
    serieGuide = endpointData.serieGuia
    endpointData.nroGuia = str(int(nroGuide) + 1)
    endpointData.save()
    codeGuide = nroGuide
    while len(codeGuide) < 4:
        codeGuide = '0' + codeGuide
    codeGuide = f'{serieGuide}-{codeGuide}'
    deparmentDeparture='02'
    provinceDeparture='SANTA'
    districtDeparture='NUEVO CHIMBOTE'
    addressDeparture='Mza. J4 Lote. 39'
    ubigeoDeparture='021809'
    razonSocialTranporter=''
    rucTransporter=''
    vehiclePlate=''
    dniDriver=''
    licenceDriver=''
    nameDriver=''
    extraWeight='0'
    ubigeoClient=''
    stateTeFacturo=''
    guideSystem.objects.create(
        asociatedQuotation=quotationObject,
        dateGuide=dateGuide,
        dateGivenGoods=dateGivenGoods,
        codeGuide=codeGuide,
        nroGuide=nroGuide,
        stateGuide='GENERADA',
        commentGuide=commentGuide,
        ubigeoClient=ubigeoClient,
        deparmentDeparture=deparmentDeparture,
        provinceDeparture=provinceDeparture,
        districtDeparture=districtDeparture,
        addressDeparture=addressDeparture,
        ubigeoDeparture=ubigeoDeparture,
        razonSocialTranporter=razonSocialTranporter,
        rucTransporter=rucTransporter,
        vehiclePlate=vehiclePlate,
        dniDriver=dniDriver,
        licenceDriver=licenceDriver,
        nameDriver=nameDriver,
        extraWeight=extraWeight,
        stateTeFacturo=stateTeFacturo,
        endpointGuide=endpointData
    )
    return HttpResponseRedirect(reverse('salesMetalprotec:guidesMetalprotec'))

def updateQuotation(request):
    if request.method == 'POST':
        quotationInfo = json.load(request)
        quotationIdInfo = quotationInfo.get('quotationIdInfo')
        productsData = quotationInfo.get('productsData')
        servicesData = quotationInfo.get('servicesData')
        clientData = quotationInfo.get('clientData')
        sellerData = quotationInfo.get('sellerData')
        quotationData = quotationInfo.get('quotationData')
        documentOptions = quotationInfo.get('documentOptions')

        editQuotationItem = quotationSystem.objects.get(id=quotationIdInfo)
        editQuotationItem.currencyQuotation=quotationData.get('currencyQuotation')
        editQuotationItem.typeQuotation=quotationData.get('typeItems')
        editQuotationItem.paymentQuotation = quotationData.get('paymentQuotation')
        editQuotationItem.dateQuotation = parse(quotationData.get('dateQuotation'))
        editQuotationItem.showDiscount = documentOptions.get('showDiscount')
        editQuotationItem.showUnitPrice = documentOptions.get('showUnitPrice')
        editQuotationItem.showSellPrice = documentOptions.get('showSellPrice')
        editQuotationItem.relatedDocumentQuotation = quotationData.get('relatedDocumentQuotation')
        editQuotationItem.commentQuotation = quotationData.get('commentQuotation')
        editQuotationItem.quotesQuotation = quotationData.get('quotesQuotation')
        editQuotationItem.expirationCredit = quotationData.get('expirationCredit')
        editQuotationItem.expirationQuotation = quotationData.get('expirationQuotation')
        editQuotationItem.erBuy = quotationData.get('erBuy')
        editQuotationItem.erSel = quotationData.get('erSel')
        editQuotationItem.save()

        clientQuotation = editQuotationItem.quotationclientdata
        clientQuotation.dataClientQuotation[1] = clientData.get('identificationClient')
        clientQuotation.dataClientQuotation[2] = clientData.get('documentClient')
        clientQuotation.dataClientQuotation[3] = clientData.get('typeClient')
        clientQuotation.dataClientQuotation[4] = clientData.get('emailClient')
        clientQuotation.dataClientQuotation[5] = clientData.get('contactClient')
        clientQuotation.dataClientQuotation[6] = clientData.get('phoneClient')
        clientQuotation.dataClientQuotation[7] = clientData.get('legalAddressClient')
        clientQuotation.dataClientQuotation[8] = clientData.get('deliveryAddress')
        clientQuotation.save()

        sellerQuotation = editQuotationItem.quotationsellerdata
        sellerQuotation.dataUserQuotation[1]=sellerData.get('nameSeller')
        sellerQuotation.dataUserQuotation[2]=sellerData.get('codeSeller')
        sellerQuotation.dataUserQuotation[3]=sellerData.get('phoneSeller')
        sellerQuotation.save()
        editQuotationItem.save()

        editQuotationItem.quotationproductdata_set.all().delete()
        editQuotationItem.quotationservicedata_set.all().delete()

        for productInfo in productsData:
            asociatedProduct = productSystem.objects.get(id=productInfo[0])
            productInfo.append(asociatedProduct.weightProduct)
            quotationProductData.objects.create(
                asociatedQuotation=editQuotationItem,
                asociatedProduct=asociatedProduct,
                dataProductQuotation=productInfo,
            )

        for serviceInfo in servicesData:
            quotationServiceData.objects.create(
                asociatedQuotation=editQuotationItem,
                asociatedService=serviceSystem.objects.get(id=serviceInfo[0]),
                dataServiceQuotation=serviceInfo,
            )
        editQuotationItem.save()
        return JsonResponse({
            'metalprotec':'200',
        })

def updateGuide(request):
    if request.method == 'POST':
        guideInfo = json.load(request)
        idGuideInfo = guideInfo.get('idGuideInfo')
        legalAddressClient = guideInfo.get('legalAddressClient')
        deliveryAddress = guideInfo.get('deliveryAddress')
        updatedWeights = guideInfo.get('updatedWeights')
        driverData = guideInfo.get('driverData')
        transporterData = guideInfo.get('transporterData')
        guideData = guideInfo.get('guideData')
        departureData = guideInfo.get('departureData')
        guideObject = guideSystem.objects.get(id=idGuideInfo)
        clientInfoData = guideObject.asociatedQuotation.quotationclientdata
        clientInfoData.dataClientQuotation[7] = legalAddressClient
        clientInfoData.dataClientQuotation[8] = deliveryAddress
        clientInfoData.save()
        guideObject.commentGuide = guideData.get('commentGuide')
        guideObject.dateGuide = parse(guideData.get('dateGuide'))
        guideObject.dateGivenGoods = parse(guideData.get('dateGivenGoods'))
        guideObject.purposeTransportation = guideData.get('purposeTransportation')
        guideObject.modeTransportation = guideData.get('modeTransportation')
        guideObject.extraWeight = guideData.get('extraWeight')
        guideObject.ubigeoClient = guideData.get('ubigeoClient')
        guideObject.deparmentDeparture = departureData.get('deparmentDeparture')
        guideObject.provinceDeparture = departureData.get('provinceDeparture')
        guideObject.districtDeparture = departureData.get('districtDeparture')
        guideObject.addressDeparture = departureData.get('addressDeparture')
        guideObject.ubigeoDeparture = departureData.get('ubigeoDeparture')
        guideObject.razonSocialTranporter = transporterData.get('razonSocialTranporter')
        guideObject.rucTransporter = transporterData.get('rucTransporter')
        guideObject.vehiclePlate = driverData.get('vehiclePlate')
        guideObject.dniDriver = driverData.get('dniDriver')
        guideObject.licenceDriver = driverData.get('licenceDriver')
        guideObject.nameDriver = driverData.get('nameDriver')
        guideObject.save()
        for weightInfo in updatedWeights:
            productObject = quotationProductData.objects.get(id=weightInfo[0])
            productObject.dataProductQuotation[10] = weightInfo[1]
            productObject.save()
        return JsonResponse({
            'metalprotec':'200',
        })

def updateBill(request):
    if request.method == 'POST':
        billInfo = json.load(request)
        
        idBillInfo = billInfo.get('idBillInfo')
        typeItemsBill = billInfo.get('typeItemsBill')
        originBill = billInfo.get('originBill')
        legalAddressClient = billInfo.get('legalAddressClient')
        deliveryAddress = billInfo.get('deliveryAddress')

        billInfoData = billInfo.get('billData')

        dateBill = billInfoData.get('dateBill')
        relatedDocumentBill = billInfoData.get('relatedDocumentBill')
        currencyBill = billInfoData.get('currencyBill')
        erBuy = billInfoData.get('erBuy')
        erSel = billInfoData.get('erSel')
        commentBill = billInfoData.get('commentBill')
        paymentQuotation = billInfoData.get('paymentQuotation')

        updatedpvnIGV = billInfo.get('updatedpvnIGV')
        updatedDiscount = billInfo.get('updatedDiscount')
        updatedQuantity = billInfo.get('updatedQuantity')
        updatedFree = billInfo.get('updatedFree')
        nroQuotes = billInfo.get('nroQuotes')
        dateQuoteBill = billInfo.get('dateQuoteBill')


        billObject = billSystem.objects.get(id=idBillInfo)

        billObject.commentBill = commentBill
        billObject.dateBill = parse(dateBill)
        billObject.relatedDocumentBill = relatedDocumentBill
        billObject.erBuy = erBuy
        billObject.erSel = erSel
        billObject.currencyBill = currencyBill
        billObject.dateQuotesBill = dateQuoteBill

        if typeItemsBill == 'PRODUCTOS':
            if originBill == 'GUIDE':
                for guideObject in billObject.guidesystem_set.all():
                    clientInfoData = guideObject.asociatedQuotation.quotationclientdata
                    clientInfoData.dataClientQuotation[7] = legalAddressClient
                    clientInfoData.dataClientQuotation[8] = deliveryAddress
                    clientInfoData.save()
            else:
                clientInfoData = billObject.asociatedQuotation.quotationclientdata
                clientInfoData.dataClientQuotation[7] = legalAddressClient
                clientInfoData.dataClientQuotation[8] = deliveryAddress
                clientInfoData.save()

            i = 0
            while i < len(updatedpvnIGV):
                productInfo = quotationProductData.objects.get(id=updatedpvnIGV[i][0])
                productInfo.dataProductQuotation[6] = updatedpvnIGV[i][1]
                productInfo.dataProductQuotation[7] = updatedDiscount[i][1]
                productInfo.dataProductQuotation[8] = updatedQuantity[i][1]
                productInfo.dataProductQuotation[9] = updatedFree[i][1]
                productInfo.save()
                i = i + 1
        else:
            clientInfoData = billObject.asociatedQuotation.quotationclientdata
            clientInfoData.dataClientQuotation[7] = legalAddressClient
            clientInfoData.dataClientQuotation[8] = deliveryAddress
            clientInfoData.save()

            i = 0
            while i < len(updatedpvnIGV):
                serviceInfo = quotationServiceData.objects.get(id=updatedpvnIGV[i][0])
                serviceInfo.dataServiceQuotation[4] = updatedpvnIGV[i][1]
                serviceInfo.dataServiceQuotation[5] = updatedDiscount[i][1]
                serviceInfo.save()
                i = i + 1
        billObject.save()
        return JsonResponse({
            'metalprotec':'200',
        })

def updateInvoice(request):
    if request.method == 'POST':
        invoiceInfo = json.load(request)
        idInvoiceInfo = invoiceInfo.get('idInvoiceInfo')
        typeItemsInvoice = invoiceInfo.get('typeItemsInvoice')
        originInvoice = invoiceInfo.get('originInvoice')
        legalAddressClient = invoiceInfo.get('legalAddressClient')
        deliveryAddress = invoiceInfo.get('deliveryAddress')

        invoiceInfoData = invoiceInfo.get('invoiceData')
        
        dateInvoice = invoiceInfoData.get('dateInvoice')
        relatedDocumentInvoice = invoiceInfoData.get('relatedDocumentInvoice')
        currencyInvoice = invoiceInfoData.get('currencyInvoice')
        erBuy = invoiceInfoData.get('erBuy')
        erSel = invoiceInfoData.get('erSel')
        commentInvoice = invoiceInfoData.get('commentInvoice')
        paymentQuotation = invoiceInfoData.get('paymentQuotation')

        updatedpvnIGV = invoiceInfo.get('updatedpvnIGV')
        updatedDiscount = invoiceInfo.get('updatedDiscount')
        updatedQuantity = invoiceInfo.get('updatedQuantity')
        updatedFree = invoiceInfo.get('updatedFree')
        nroQuotes = invoiceInfo.get('nroQuotes')
        dateQuoteInvoice = invoiceInfo.get('dateQuoteInvoice')


        invoiceObject = invoiceSystem.objects.get(id=idInvoiceInfo)

        invoiceObject.commentInvoice = commentInvoice
        invoiceObject.dateInvoice = parse(dateInvoice)
        invoiceObject.relatedDocumentInvoice = relatedDocumentInvoice
        invoiceObject.erBuy = erBuy
        invoiceObject.erSel = erSel
        invoiceObject.currencyInvoice = currencyInvoice
        invoiceObject.dateQuotesInvoice = dateQuoteInvoice

        if typeItemsInvoice == 'PRODUCTOS':
            if originInvoice == 'GUIDE':
                for guideObject in invoiceObject.guidesystem_set.all():
                    clientInfoData = guideObject.asociatedQuotation.quotationclientdata
                    clientInfoData.dataClientQuotation[7] = legalAddressClient
                    clientInfoData.dataClientQuotation[8] = deliveryAddress
                    clientInfoData.save()
            else:
                clientInfoData = invoiceObject.asociatedQuotation.quotationclientdata
                clientInfoData.dataClientQuotation[7] = legalAddressClient
                clientInfoData.dataClientQuotation[8] = deliveryAddress
                clientInfoData.save()

            i = 0
            while i < len(updatedpvnIGV):
                productInfo = quotationProductData.objects.get(id=updatedpvnIGV[i][0])
                productInfo.dataProductQuotation[6] = updatedpvnIGV[i][1]
                productInfo.dataProductQuotation[7] = updatedDiscount[i][1]
                productInfo.dataProductQuotation[8] = updatedQuantity[i][1]
                productInfo.dataProductQuotation[9] = updatedFree[i][1]
                productInfo.save()
                i = i + 1
        else:
            clientInfoData = invoiceObject.asociatedQuotation.quotationclientdata
            clientInfoData.dataClientQuotation[7] = legalAddressClient
            clientInfoData.dataClientQuotation[8] = deliveryAddress
            clientInfoData.save()

            i = 0
            while i < len(updatedpvnIGV):
                serviceInfo = quotationServiceData.objects.get(id=updatedpvnIGV[i][0])
                serviceInfo.dataServiceQuotation[4] = updatedpvnIGV[i][1]
                serviceInfo.dataServiceQuotation[5] = updatedDiscount[i][1]
                serviceInfo.save()
                i = i + 1
        invoiceObject.save()
        return JsonResponse({
            'metalprotec':'200',
        })

def createBillFromGuide(request,idGuide):
    typeItemsBill = 'PRODUCTOS'
    originBill = 'GUIDE'
    guideObject = guideSystem.objects.get(id=idGuide)
    dateBill = datetime.datetime.today()
    endpointData = endpointSystem.objects.get(id=request.user.extendeduser.endpointUser.id)
    commentBill=guideObject.commentGuide
    relatedDocumentBill=guideObject.asociatedQuotation.relatedDocumentQuotation
    stateTeFacturo = ''
    numberBill = endpointData.nroFactura
    serieBill = endpointData.serieFactura
    endpointData.nroFactura = str(int(numberBill) + 1)
    endpointData.save()
    codeBill = numberBill
    while len(codeBill) < 4:
        codeBill = '0' + codeBill
    codeBill = f'{serieBill}-{codeBill}'
    dateQuotesBill = []
    if guideObject.asociatedQuotation.paymentQuotation == 'CONTADO':
        dateQuotesBill = []
    else:
        for numberData in range(int(guideObject.asociatedQuotation.quotesQuotation)):
            dateQuotesBill.append('2023-01-01')
    billInfo = billSystem.objects.create(
        stateBill='GENERADA',
        dateBill=dateBill,
        erBuy=guideObject.asociatedQuotation.erBuy,
        erSel=guideObject.asociatedQuotation.erSel,
        currencyBill=guideObject.asociatedQuotation.currencyQuotation,
        nroBill = numberBill,
        codeBill = codeBill,
        dateQuotesBill = dateQuotesBill,
        commentBill=commentBill,
        relatedDocumentBill=relatedDocumentBill,
        stateTeFacturo=stateTeFacturo,
        typeItemsBill=typeItemsBill,
        originBill=originBill,
        endpointBill=request.user.extendeduser.endpointUser,
    )
    guideObject.asociatedBill = billInfo
    guideObject.stateGuide = 'EMITIDA'
    guideObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:billsMetalprotec'))

def createInvoiceFromGuide(request,idGuide):
    typeItemsInvoice = 'PRODUCTOS'
    originInvoice = 'GUIDE'
    guideObject = guideSystem.objects.get(id=idGuide)
    dateInvoice = datetime.datetime.today()
    endpointData = endpointSystem.objects.get(id=request.user.extendeduser.endpointUser.id)
    commentInvoice=guideObject.commentGuide
    relatedDocumentInvoice=guideObject.asociatedQuotation.relatedDocumentQuotation
    stateTeFacturo = ''
    numberInvoice = endpointData.nroBoleta
    serieInvoice = endpointData.serieBoleta
    endpointData.nroBoleta = str(int(numberInvoice) + 1)
    endpointData.save()
    codeInvoice = numberInvoice
    while len(codeInvoice) < 4:
        codeInvoice = '0' + codeInvoice
    codeInvoice = f'{serieInvoice}-{codeInvoice}'
    dateQuotesInvoice = []
    if guideObject.asociatedQuotation.paymentQuotation == 'CONTADO':
        dateQuotesInvoice = []
    else:
        for numberData in range(int(guideObject.asociatedQuotation.quotesQuotation)):
            dateQuotesInvoice.append('2023-01-01')
    invoiceInfo = invoiceSystem.objects.create(
        stateInvoice='GENERADA',
        dateInvoice=dateInvoice,
        erBuy=guideObject.asociatedQuotation.erBuy,
        erSel=guideObject.asociatedQuotation.erSel,
        currencyInvoice=guideObject.asociatedQuotation.currencyQuotation,
        nroInvoice = numberInvoice,
        codeInvoice = codeInvoice,
        dateQuotesInvoice = dateQuotesInvoice,
        commentInvoice=commentInvoice,
        relatedDocumentInvoice=relatedDocumentInvoice,
        stateTeFacturo=stateTeFacturo,
        typeItemsInvoice=typeItemsInvoice,
        originInvoice=originInvoice,
        endpointInvoice=request.user.extendeduser.endpointUser,
    )
    print(invoiceInfo)
    guideObject.asociatedInvoice = invoiceInfo
    guideObject.stateGuide = 'EMITIDA'
    guideObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:invoicesMetalprotec'))



def createBillFromQuotation(request, idQuotation):
    asociatedQuotation = quotationSystem.objects.get(id=idQuotation)
    asociatedQuotation.stateQuotation = 'EMITIDA'
    asociatedQuotation.save()
    commentBill = asociatedQuotation.commentQuotation
    dateBill = datetime.datetime.today()
    relatedDocumentBill = asociatedQuotation.relatedDocumentQuotation
    erBuy = asociatedQuotation.erBuy
    erSel = asociatedQuotation.erSel
    currencyBill = asociatedQuotation.currencyQuotation
    stateBill = 'GENERADA'
    stateTeFacturo = ''
    typeItemsBill = asociatedQuotation.typeQuotation
    originBill = 'QUOTATION'
    endpointBill = request.user.extendeduser.endpointUser
    numberBill = endpointBill.nroFactura
    serieBill = endpointBill.serieFactura
    endpointBill.nroFactura = str(int(numberBill) + 1)
    endpointBill.save()
    codeBill = numberBill
    while len(codeBill) < 4:
        codeBill = '0' + codeBill
    codeBill = f'{serieBill}-{codeBill}'
    dateQuotesBill = []
    if asociatedQuotation.paymentQuotation == 'CONTADO':
        dateQuotesBill = []
    else:
        for numberData in range(int(asociatedQuotation.quotesQuotation)):
            dateQuotesBill.append('2023-01-01')
    billSystem.objects.create(
        commentBill=commentBill,
        dateBill=dateBill,
        relatedDocumentBill=relatedDocumentBill,
        dateQuotesBill=dateQuotesBill,
        erBuy=erBuy,
        erSel=erSel,
        currencyBill=currencyBill,
        stateBill=stateBill,
        stateTeFacturo=stateTeFacturo,
        nroBill=numberBill,
        typeItemsBill=typeItemsBill,
        originBill=originBill,
        asociatedQuotation=asociatedQuotation,
        codeBill=codeBill,
        endpointBill=endpointBill
    )
    return HttpResponseRedirect(reverse('salesMetalprotec:billsMetalprotec'))

def createInvoiceFromQuotation(request,idQuotation):
    asociatedQuotation = quotationSystem.objects.get(id=idQuotation)
    asociatedQuotation.stateQuotation = 'EMITIDA'
    asociatedQuotation.save()
    commentInvoice = asociatedQuotation.commentQuotation
    dateInvoice = datetime.datetime.today()
    relatedDocumentInvoice = asociatedQuotation.relatedDocumentQuotation
    erBuy = asociatedQuotation.erBuy
    erSel = asociatedQuotation.erSel
    currencyInvoice = asociatedQuotation.currencyQuotation
    stateInvoice = 'GENERADA'
    stateTeFacturo = ''
    typeItemsInvoice = asociatedQuotation.typeQuotation
    originInvoice = 'QUOTATION'
    endpointInvoice = request.user.extendeduser.endpointUser
    numberInvoice = endpointInvoice.nroBoleta
    serieInvoice = endpointInvoice.serieBoleta
    endpointInvoice.nroBoleta = str(int(numberInvoice) + 1)
    endpointInvoice.save()
    codeInvoice = numberInvoice
    while len(codeInvoice) < 4:
        codeInvoice = '0' + codeInvoice
    codeInvoice = f'{serieInvoice}-{codeInvoice}'
    dateQuotesInvoice = []
    if asociatedQuotation.paymentQuotation == 'CONTADO':
        dateQuotesInvoice = []
    else:
        for numberData in range(int(asociatedQuotation.quotesQuotation)):
            dateQuotesInvoice.append('2023-01-01')
    invoiceSystem.objects.create(
        commentInvoice=commentInvoice,
        dateInvoice=dateInvoice,
        relatedDocumentInvoice=relatedDocumentInvoice,
        dateQuotesInvoice=dateQuotesInvoice,
        erBuy=erBuy,
        erSel=erSel,
        currencyInvoice=currencyInvoice,
        stateInvoice=stateInvoice,
        stateTeFacturo=stateTeFacturo,
        nroInvoice=numberInvoice,
        typeItemsInvoice=typeItemsInvoice,
        originInvoice=originInvoice,
        asociatedQuotation=asociatedQuotation,
        codeInvoice=codeInvoice,
        endpointInvoice=endpointInvoice
    )
    return HttpResponseRedirect(reverse('salesMetalprotec:invoicesMetalprotec'))

def cancelQuotation(request,idQuotation):
    quotationObject = quotationSystem.objects.get(id=idQuotation)
    quotationObject.stateQuotation = 'ANULADA'
    quotationObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:quotationsMetalprotec'))

def downloadQuotationDolares(request,idQuotation):
    #Generacion del documento
    pdf_name = 'quotationMetalprotec.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)

    quotationItem = quotationSystem.objects.get(id=idQuotation)
    quotationItem.currencyQuotation = 'DOLARES'

    if quotationItem.typeQuotation == 'PRODUCTOS':
        totalProductsQuotation = quotationItem.quotationproductdata_set.all()
        groupQuantity = [totalProductsQuotation[x:x+40] for x in range(0,len(totalProductsQuotation),40)]
        totalGroups = len(groupQuantity)
        indicatorGroup = 0
        total_precio = Decimal(0.0000)
        while indicatorGroup < totalGroups:
            can.setStrokeColorRGB(0,0,1)
            lista_x = [410,570]
            lista_y = [750,810]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,790,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,775,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(quotationItem.numberQuotation)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,760,str(quotationItem.endpointQuotation.serieCoti) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./static/salesMetalprotec/images/logoNuevo.png',25,745,width=90,height=60,mask='auto')
            
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,730,'METALPROTEC S.A.C')
            can.drawString(25,720,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,700,'OFICINA :')
            can.drawString(25,690,'LOCAL COMERCIAL :')
            can.drawString(25,680,'TELÉFONO :')
            can.drawString(450,680,f'FECHA : {quotationItem.dateQuotation.strftime("%d-%m-%Y")}')
            can.setFont('Helvetica',7)
            can.drawString(60,700,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(100,690,'AV. JOSE PARDO 2721 URB. MIRAFLORES ALTO - CHIMBOTE')
            can.drawString(70,680,'(043) 282752')


            dato_imprimir = 'Pagina ' + str(indicatorGroup + 1) + ' de ' + str(totalGroups)
            can.drawString(25,815,dato_imprimir)
            
            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,660,'SEÑORES :')
            can.setFont('Helvetica',7)
            can.drawString(120,660,str(quotationItem.quotationclientdata.dataClientQuotation[1]))
            
            if quotationItem.quotationclientdata.dataClientQuotation[3] == 'EMPRESA':
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'RUC :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(quotationItem.quotationclientdata.dataClientQuotation[2]))
            else:
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'DNI :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(quotationItem.quotationclientdata.dataClientQuotation[2]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,640,'DIRECCIÓN :')
            can.setFont('Helvetica',7)
            can.drawString(120,640,str(quotationItem.quotationclientdata.dataClientQuotation[7]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,630,'FORMA DE PAGO :')
            can.setFont('Helvetica',7)
            if quotationItem.paymentQuotation == 'CONTADO':
                can.drawString(120,630,str(quotationItem.paymentQuotation))
            if quotationItem.paymentQuotation == 'CREDITO':
                can.drawString(120,630,str(quotationItem.paymentQuotation) + ' ' + str(quotationItem.expirationCredit))
            
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,620,'VALIDEZ :')
            can.setFont('Helvetica',7)
            can.drawString(120,620,str(quotationItem.expirationQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,610,'ENTREGA :')
            can.setFont('Helvetica',7)
            can.drawString(120,610,"SEGUN STOCK")

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,600,'MONEDA :')
            can.setFont('Helvetica',7)
            can.drawString(120,600,str(quotationItem.currencyQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,590,'PESO APROX :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,580,'DOCUMENTO :')
            can.setFont('Helvetica',7)
            can.drawString(120,580,str(quotationItem.relatedDocumentQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,570,'OBRA :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,560,'OBSERVACIONES :')
            can.setFont('Helvetica',6)
            can.drawString(120,560,str(quotationItem.commentQuotation))

            #Linea de separacion con los datos del vendedor
            can.line(25,550,580,550)

            #Datos del vendedor
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,540,'VENDEDOR :')
            can.setFont('Helvetica',7)
            can.drawString(120,540,str(quotationItem.quotationsellerdata.dataUserQuotation[1]))
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,530,'CELULAR :')
            can.setFont('Helvetica',7)
            can.drawString(120,530,str(quotationItem.quotationsellerdata.dataUserQuotation[3]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,520,'EMAIL:')
            can.setFont('Helvetica',7)
            can.drawString(120,520,str(quotationItem.quotationsellerdata.asociatedUser.email))

            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [500,515]
            can.setFillColorRGB(0,0,1)
            can.rect(25,500,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,120,310,360,410,460,530]
            lista_y = [500,515]
            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Cant.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("{:.2f}".format(round(float(productInfo.dataProductQuotation[8]),2))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,productInfo.dataProductQuotation[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,productInfo.dataProductQuotation[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,productInfo.dataProductQuotation[3])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            condicion_imprimir = quotationItem.showUnitPrice + quotationItem.showSellPrice + quotationItem.showDiscount
            if condicion_imprimir == 'ONOFFOFF':
                lista_x[4] = 360
            
            if condicion_imprimir == 'ONOFFON':
                lista_x[5] = 360

            if condicion_imprimir == 'OFFOFFON':
                lista_x[6] = 360
            
            if condicion_imprimir == 'ONONOFF':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == 'ONOFFON':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == 'OFFONON':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == 'ONONON':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480

            if quotationItem.showUnitPrice == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for productInfo in groupQuantity[indicatorGroup]:
                    if quotationItem.currencyQuotation == 'SOLES':
                        if productInfo.dataProductQuotation[5] == 'DOLARES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        if productInfo.dataProductQuotation[5] == 'SOLES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                    if quotationItem.currencyQuotation == 'DOLARES':
                        if productInfo.dataProductQuotation[5] == 'SOLES':
                            vu_producto = (Decimal(productInfo.dataProductQuotation[6])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        if productInfo.dataProductQuotation[5] == 'DOLARES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                    if productInfo.dataProductQuotation[9] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_producto)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if quotationItem.showSellPrice == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for productInfo in groupQuantity[indicatorGroup]:
                    if quotationItem.currencyQuotation == 'SOLES':
                        if productInfo.dataProductQuotation[5] == 'DOLARES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        if productInfo.dataProductQuotation[5] == 'SOLES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                    if quotationItem.currencyQuotation == 'DOLARES':
                        if productInfo.dataProductQuotation[5] == 'SOLES':
                            vu_producto = (Decimal(productInfo.dataProductQuotation[6])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        if productInfo.dataProductQuotation[5] == 'DOLARES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                    if productInfo.dataProductQuotation[9] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_producto*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if quotationItem.showDiscount == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for productInfo in groupQuantity[indicatorGroup]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(productInfo.dataProductQuotation[7]) + ' %')
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]


            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                if quotationItem.currencyQuotation == 'SOLES':
                    if productInfo.dataProductQuotation[5] == 'DOLARES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                    if productInfo.dataProductQuotation[5] == 'SOLES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                if quotationItem.currencyQuotation == 'DOLARES':
                    if productInfo.dataProductQuotation[5] == 'SOLES':
                        v_producto = (Decimal(productInfo.dataProductQuotation[6])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                    if productInfo.dataProductQuotation[5] == 'DOLARES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                #v_producto = round(v_producto,2)
                if productInfo.dataProductQuotation[9] == '1':
                    v_producto = Decimal(0.00)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(v_producto))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(v_producto)
            
            
            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])

            #Logo de las empresas
            can.drawImage('./static/salesMetalprotec/images/mega.png',20,60,width=90,height=60,mask='auto')
            can.drawImage('./static/salesMetalprotec/images/hodelpe.png',470,65,width=100,height=50,mask='auto')

            #Linea inicio de separacion
            can.line(25,60,580,60)

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,50,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,40,'Cta Cte Soles: 000 9496505')
            can.drawString(25,30,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(230,50,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(230,40,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(230,30,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(420,50,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(420,40,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(420,30,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)

            indicatorGroup = indicatorGroup + 1
            if totalGroups > indicatorGroup:
                can.showPage()

        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo del valor en soles
        total_soles = Decimal(0.0000)
        for productInfo in totalProductsQuotation:
            if productInfo.dataProductQuotation[5] == 'DOLARES':
                v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
            if productInfo.dataProductQuotation[5] == 'SOLES':
                v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
            if productInfo.dataProductQuotation[9] == '1':
                v_producto = Decimal(0.00)
            total_soles = Decimal(total_soles) + Decimal(v_producto)
        final_soles = Decimal('%.2f' % total_soles)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_soles))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        
        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()

    if quotationItem.typeQuotation == 'SERVICIOS':
    
        totalServicesQuotation = quotationItem.quotationservicedata_set.all()
        groupQuantity = [totalServicesQuotation[x:x+40] for x in range(0,len(totalServicesQuotation),40)]
        totalGroups = len(groupQuantity)
        indicatorGroup = 0
        total_precio = Decimal(0.0000)

        while indicatorGroup < totalGroups:
            can.setStrokeColorRGB(0,0,1)
            lista_x = [410,570]
            lista_y = [750,810]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,790,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,775,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(quotationItem.numberQuotation)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,760,str(quotationItem.endpointQuotation.serieCoti) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./static/salesMetalprotec/images/logoNuevo.png',25,745,width=90,height=60,mask='auto')
            
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,730,'METALPROTEC S.A.C')
            can.drawString(25,720,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,700,'OFICINA :')
            can.drawString(25,690,'LOCAL COMERCIAL :')
            can.drawString(25,680,'TELÉFONO :')
            can.drawString(450,680,f'FECHA : {quotationItem.dateQuotation.strftime("%d-%m-%Y")}')
            can.setFont('Helvetica',7)
            can.drawString(60,700,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(100,690,'AV. JOSE PARDO 2721 URB. MIRAFLORES ALTO - CHIMBOTE')
            can.drawString(70,680,'(043) 282752')


            dato_imprimir = 'Pagina ' + str(indicatorGroup + 1) + ' de ' + str(totalGroups)
            can.drawString(25,815,dato_imprimir)
            
            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,660,'SEÑORES :')
            can.setFont('Helvetica',7)
            can.drawString(120,660,str(quotationItem.quotationclientdata.dataClientQuotation[1]))
            
            if quotationItem.quotationclientdata.dataClientQuotation[3] == 'EMPRESA':
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'RUC :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(quotationItem.quotationclientdata.dataClientQuotation[2]))
            else:
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'DNI :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(quotationItem.quotationclientdata.dataClientQuotation[2]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,640,'DIRECCIÓN :')
            can.setFont('Helvetica',7)
            can.drawString(120,640,str(quotationItem.quotationclientdata.dataClientQuotation[7]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,630,'FORMA DE PAGO :')
            can.setFont('Helvetica',7)
            if quotationItem.paymentQuotation == 'CONTADO':
                can.drawString(120,630,str(quotationItem.paymentQuotation))
            if quotationItem.paymentQuotation == 'CREDITO':
                can.drawString(120,630,str(quotationItem.paymentQuotation) + ' ' + str(quotationItem.expirationCredit))
            
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,620,'VALIDEZ :')
            can.setFont('Helvetica',7)
            can.drawString(120,620,str(quotationItem.expirationQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,610,'ENTREGA :')
            can.setFont('Helvetica',7)
            can.drawString(120,610,"SEGUN STOCK")

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,600,'MONEDA :')
            can.setFont('Helvetica',7)
            can.drawString(120,600,str(quotationItem.currencyQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,590,'PESO APROX :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,580,'DOCUMENTO :')
            can.setFont('Helvetica',7)
            can.drawString(120,580,str(quotationItem.relatedDocumentQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,570,'OBRA :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,560,'OBSERVACIONES :')
            can.setFont('Helvetica',6)
            can.drawString(120,560,str(quotationItem.commentQuotation))

            #Linea de separacion con los datos del vendedor
            can.line(25,550,580,550)

            #Datos del vendedor
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,540,'VENDEDOR :')
            can.setFont('Helvetica',7)
            can.drawString(120,540,str(quotationItem.quotationsellerdata.dataUserQuotation[1]))
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,530,'CELULAR :')
            can.setFont('Helvetica',7)
            can.drawString(120,530,str(quotationItem.quotationsellerdata.dataUserQuotation[3]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,520,'EMAIL:')
            can.setFont('Helvetica',7)
            can.drawString(120,520,str(quotationItem.quotationsellerdata.asociatedUser.email))


            #Aqui se ponen las cabeceras

            can.setStrokeColorRGB(0,0,1)
            can.setFillColorRGB(0,0,0)
            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [500,515]
            can.setFillColorRGB(0,0,1)
            can.rect(25,500,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,120,310,360,410,460,530]
            lista_y = [500,515]

            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Item.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("1"))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,'Servicio')
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,serviceInfo.dataServiceQuotation[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,serviceInfo.dataServiceQuotation[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            condicion_imprimir = quotationItem.showUnitPrice + quotationItem.showSellPrice + quotationItem.showDiscount
            if condicion_imprimir == 'ONOFFOFF':
                lista_x[4] = 360
            
            if condicion_imprimir == 'ONOFFON':
                lista_x[5] = 360

            if condicion_imprimir == 'OFFOFFON':
                lista_x[6] = 360
            
            if condicion_imprimir == 'ONONOFF':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == 'ONOFFON':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == 'OFFONON':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == 'ONONON':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480


            if quotationItem.showUnitPrice == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for serviceInfo in groupQuantity[indicatorGroup]:
                    if quotationItem.currencyQuotation == 'SOLES':
                        if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                        if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                    if quotationItem.currencyQuotation == 'DOLARES':
                        if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                            vu_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                        if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_servicio)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if quotationItem.showSellPrice == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for serviceInfo in groupQuantity[indicatorGroup]:
                    if quotationItem.currencyQuotation == 'SOLES':
                        if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                        if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                    if quotationItem.currencyQuotation == 'DOLARES':
                        if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                            vu_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                        if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_servicio*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if quotationItem.showDiscount == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for serviceInfo in groupQuantity[indicatorGroup]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(serviceInfo.dataServiceQuotation[5]) + ' %')
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                if quotationItem.currencyQuotation == 'SOLES':
                    if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                        vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                    if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                        vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                if quotationItem.currencyQuotation == 'DOLARES':
                    if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                        vu_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                    if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                        vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                #v_producto = round(v_producto,2)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(vu_servicio))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(vu_servicio)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])

            #Logo de las empresas
            can.drawImage('./static/salesMetalprotec/images/mega.png',20,60,width=90,height=60,mask='auto')
            can.drawImage('./static/salesMetalprotec/images/hodelpe.png',470,65,width=100,height=50,mask='auto')

            #Linea inicio de separacion
            can.line(25,60,580,60)

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,50,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,40,'Cta Cte Soles: 000 9496505')
            can.drawString(25,30,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(230,50,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(230,40,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(230,30,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(420,50,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(420,40,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(420,30,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)

            indicatorGroup = indicatorGroup + 1
            if totalGroups > indicatorGroup:
                can.showPage()
        
        #Esta seccion solo va en la hoja final de los productos
        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo de la proforma en dolares
        total_soles = Decimal(0.0000)
        for serviceInfo in totalServicesQuotation:
            if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                v_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
            if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                v_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
            total_soles = Decimal(total_soles) + Decimal(v_servicio)
        final_soles = Decimal('%.2f' % total_soles)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(final_soles)))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()

    nombre_doc = str(quotationItem.codeQuotation) + '.pdf'
    response = HttpResponse(open('quotationMetalprotec.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

def downloadQuotationSoles(request,idQuotation):
    #Generacion del documento
    pdf_name = 'quotationMetalprotec.pdf'
    can = canvas.Canvas(pdf_name,pagesize=A4)

    quotationItem = quotationSystem.objects.get(id=idQuotation)
    quotationItem.currencyQuotation = 'SOLES'

    if quotationItem.typeQuotation == 'PRODUCTOS':
        totalProductsQuotation = quotationItem.quotationproductdata_set.all()
        groupQuantity = [totalProductsQuotation[x:x+40] for x in range(0,len(totalProductsQuotation),40)]
        totalGroups = len(groupQuantity)
        indicatorGroup = 0
        total_precio = Decimal(0.0000)
        while indicatorGroup < totalGroups:
            can.setStrokeColorRGB(0,0,1)
            lista_x = [410,570]
            lista_y = [750,810]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,790,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,775,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(quotationItem.numberQuotation)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,760,str(quotationItem.endpointQuotation.serieCoti) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./static/salesMetalprotec/images/logoNuevo.png',25,745,width=90,height=60,mask='auto')
            
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,730,'METALPROTEC S.A.C')
            can.drawString(25,720,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,700,'OFICINA :')
            can.drawString(25,690,'LOCAL COMERCIAL :')
            can.drawString(25,680,'TELÉFONO :')
            can.drawString(450,680,f'FECHA : {quotationItem.dateQuotation.strftime("%d-%m-%Y")}')
            can.setFont('Helvetica',7)
            can.drawString(60,700,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(100,690,'AV. JOSE PARDO 2721 URB. MIRAFLORES ALTO - CHIMBOTE')
            can.drawString(70,680,'(043) 282752')


            dato_imprimir = 'Pagina ' + str(indicatorGroup + 1) + ' de ' + str(totalGroups)
            can.drawString(25,815,dato_imprimir)
            
            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,660,'SEÑORES :')
            can.setFont('Helvetica',7)
            can.drawString(120,660,str(quotationItem.quotationclientdata.dataClientQuotation[1]))
            
            if quotationItem.quotationclientdata.dataClientQuotation[3] == 'EMPRESA':
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'RUC :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(quotationItem.quotationclientdata.dataClientQuotation[2]))
            else:
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'DNI :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(quotationItem.quotationclientdata.dataClientQuotation[2]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,640,'DIRECCIÓN :')
            can.setFont('Helvetica',7)
            can.drawString(120,640,str(quotationItem.quotationclientdata.dataClientQuotation[7]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,630,'FORMA DE PAGO :')
            can.setFont('Helvetica',7)
            if quotationItem.paymentQuotation == 'CONTADO':
                can.drawString(120,630,str(quotationItem.paymentQuotation))
            if quotationItem.paymentQuotation == 'CREDITO':
                can.drawString(120,630,str(quotationItem.paymentQuotation) + ' ' + str(quotationItem.expirationCredit))
            
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,620,'VALIDEZ :')
            can.setFont('Helvetica',7)
            can.drawString(120,620,str(quotationItem.expirationQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,610,'ENTREGA :')
            can.setFont('Helvetica',7)
            can.drawString(120,610,"SEGUN STOCK")

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,600,'MONEDA :')
            can.setFont('Helvetica',7)
            can.drawString(120,600,str(quotationItem.currencyQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,590,'PESO APROX :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,580,'DOCUMENTO :')
            can.setFont('Helvetica',7)
            can.drawString(120,580,str(quotationItem.relatedDocumentQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,570,'OBRA :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,560,'OBSERVACIONES :')
            can.setFont('Helvetica',6)
            can.drawString(120,560,str(quotationItem.commentQuotation))

            #Linea de separacion con los datos del vendedor
            can.line(25,550,580,550)

            #Datos del vendedor
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,540,'VENDEDOR :')
            can.setFont('Helvetica',7)
            can.drawString(120,540,str(quotationItem.quotationsellerdata.dataUserQuotation[1]))
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,530,'CELULAR :')
            can.setFont('Helvetica',7)
            can.drawString(120,530,str(quotationItem.quotationsellerdata.dataUserQuotation[3]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,520,'EMAIL:')
            can.setFont('Helvetica',7)
            can.drawString(120,520,str(quotationItem.quotationsellerdata.asociatedUser.email))

            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [500,515]
            can.setFillColorRGB(0,0,1)
            can.rect(25,500,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,120,310,360,410,460,530]
            lista_y = [500,515]
            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Cant.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("{:.2f}".format(round(float(productInfo.dataProductQuotation[8]),2))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,productInfo.dataProductQuotation[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,productInfo.dataProductQuotation[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,productInfo.dataProductQuotation[3])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            condicion_imprimir = quotationItem.showUnitPrice + quotationItem.showSellPrice + quotationItem.showDiscount
            if condicion_imprimir == 'ONOFFOFF':
                lista_x[4] = 360
            
            if condicion_imprimir == 'ONOFFON':
                lista_x[5] = 360

            if condicion_imprimir == 'OFFOFFON':
                lista_x[6] = 360
            
            if condicion_imprimir == 'ONONOFF':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == 'ONOFFON':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == 'OFFONON':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == 'ONONON':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480

            if quotationItem.showUnitPrice == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for productInfo in groupQuantity[indicatorGroup]:
                    if quotationItem.currencyQuotation == 'SOLES':
                        if productInfo.dataProductQuotation[5] == 'DOLARES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        if productInfo.dataProductQuotation[5] == 'SOLES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                    if quotationItem.currencyQuotation == 'DOLARES':
                        if productInfo.dataProductQuotation[5] == 'SOLES':
                            vu_producto = (Decimal(productInfo.dataProductQuotation[6])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        if productInfo.dataProductQuotation[5] == 'DOLARES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                    if productInfo.dataProductQuotation[9] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_producto)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if quotationItem.showSellPrice == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for productInfo in groupQuantity[indicatorGroup]:
                    if quotationItem.currencyQuotation == 'SOLES':
                        if productInfo.dataProductQuotation[5] == 'DOLARES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        if productInfo.dataProductQuotation[5] == 'SOLES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                    if quotationItem.currencyQuotation == 'DOLARES':
                        if productInfo.dataProductQuotation[5] == 'SOLES':
                            vu_producto = (Decimal(productInfo.dataProductQuotation[6])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        if productInfo.dataProductQuotation[5] == 'DOLARES':
                            vu_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                    if productInfo.dataProductQuotation[9] == '1':
                        vu_producto = Decimal(0.00)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_producto*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if quotationItem.showDiscount == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for productInfo in groupQuantity[indicatorGroup]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(productInfo.dataProductQuotation[7]) + ' %')
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]


            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for productInfo in groupQuantity[indicatorGroup]:
                if quotationItem.currencyQuotation == 'SOLES':
                    if productInfo.dataProductQuotation[5] == 'DOLARES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                    if productInfo.dataProductQuotation[5] == 'SOLES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                if quotationItem.currencyQuotation == 'DOLARES':
                    if productInfo.dataProductQuotation[5] == 'SOLES':
                        v_producto = (Decimal(productInfo.dataProductQuotation[6])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                    if productInfo.dataProductQuotation[5] == 'DOLARES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                #v_producto = round(v_producto,2)
                if productInfo.dataProductQuotation[9] == '1':
                    v_producto = Decimal(0.00)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(v_producto))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(v_producto)
            
            
            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])

            #Logo de las empresas
            can.drawImage('./static/salesMetalprotec/images/mega.png',20,60,width=90,height=60,mask='auto')
            can.drawImage('./static/salesMetalprotec/images/hodelpe.png',470,65,width=100,height=50,mask='auto')

            #Linea inicio de separacion
            can.line(25,60,580,60)

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,50,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,40,'Cta Cte Soles: 000 9496505')
            can.drawString(25,30,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(230,50,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(230,40,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(230,30,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(420,50,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(420,40,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(420,30,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)

            indicatorGroup = indicatorGroup + 1
            if totalGroups > indicatorGroup:
                can.showPage()

        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo del valor en soles
        total_dolares = Decimal(0.0000)
        for productInfo in totalProductsQuotation:
            if productInfo.dataProductQuotation[5] == 'SOLES':
                v_producto = Decimal(productInfo.dataProductQuotation[6])/Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
            if productInfo.dataProductQuotation[5] == 'DOLARES':
                v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
            if productInfo.dataProductQuotation[9] == '1':
                v_producto = Decimal(0.00)
            total_dolares = Decimal(total_dolares) + Decimal(v_producto)
        final_dolares = Decimal('%.2f' % total_dolares)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_dolares))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        
        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()

    if quotationItem.typeQuotation == 'SERVICIOS':
    
        totalServicesQuotation = quotationItem.quotationservicedata_set.all()
        groupQuantity = [totalServicesQuotation[x:x+40] for x in range(0,len(totalServicesQuotation),40)]
        totalGroups = len(groupQuantity)
        indicatorGroup = 0
        total_precio = Decimal(0.0000)

        while indicatorGroup < totalGroups:
            can.setStrokeColorRGB(0,0,1)
            lista_x = [410,570]
            lista_y = [750,810]
            can.grid(lista_x,lista_y)
            can.setFillColorRGB(0,0,0)
            can.setFont('Helvetica',12)
            can.drawString(440,790,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',12)
            can.drawString(455,775,'COTIZACION')
            can.setFont('Helvetica',12)
            numImp = str(quotationItem.numberQuotation)
            if len(numImp) < 4:
                while(len(numImp) < 4):
                    numImp = '0' + numImp
            else:
                pass
            can.drawString(460,760,str(quotationItem.endpointQuotation.serieCoti) + ' - ' + numImp)

            #Generacion del logo
            can.drawImage('./static/salesMetalprotec/images/logoNuevo.png',25,745,width=90,height=60,mask='auto')
            
            
            #Informacion del remitente
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,730,'METALPROTEC S.A.C')
            can.drawString(25,720,'RUC: 20541628631')
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,700,'OFICINA :')
            can.drawString(25,690,'LOCAL COMERCIAL :')
            can.drawString(25,680,'TELÉFONO :')
            can.drawString(450,680,f'FECHA : {quotationItem.dateQuotation.strftime("%d-%m-%Y")}')
            can.setFont('Helvetica',7)
            can.drawString(60,700,'LT 39 MZ. J4 URB. PASEO DEL MAR - ÁNCASH SANTA NUEVO CHIMBOTE')
            can.drawString(100,690,'AV. JOSE PARDO 2721 URB. MIRAFLORES ALTO - CHIMBOTE')
            can.drawString(70,680,'(043) 282752')


            dato_imprimir = 'Pagina ' + str(indicatorGroup + 1) + ' de ' + str(totalGroups)
            can.drawString(25,815,dato_imprimir)
            
            #Generacion de la linea de separacion
            can.line(25,670,580,670)

            #Generacion de los datos del cliente
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,660,'SEÑORES :')
            can.setFont('Helvetica',7)
            can.drawString(120,660,str(quotationItem.quotationclientdata.dataClientQuotation[1]))
            
            if quotationItem.quotationclientdata.dataClientQuotation[3] == 'EMPRESA':
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'RUC :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(quotationItem.quotationclientdata.dataClientQuotation[2]))
            else:
                can.setFont('Helvetica-Bold',7)
                can.drawString(25,650,'DNI :')
                can.setFont('Helvetica',7)
                can.drawString(120,650,str(quotationItem.quotationclientdata.dataClientQuotation[2]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,640,'DIRECCIÓN :')
            can.setFont('Helvetica',7)
            can.drawString(120,640,str(quotationItem.quotationclientdata.dataClientQuotation[7]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,630,'FORMA DE PAGO :')
            can.setFont('Helvetica',7)
            if quotationItem.paymentQuotation == 'CONTADO':
                can.drawString(120,630,str(quotationItem.paymentQuotation))
            if quotationItem.paymentQuotation == 'CREDITO':
                can.drawString(120,630,str(quotationItem.paymentQuotation) + ' ' + str(quotationItem.expirationCredit))
            
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,620,'VALIDEZ :')
            can.setFont('Helvetica',7)
            can.drawString(120,620,str(quotationItem.expirationQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,610,'ENTREGA :')
            can.setFont('Helvetica',7)
            can.drawString(120,610,"SEGUN STOCK")

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,600,'MONEDA :')
            can.setFont('Helvetica',7)
            can.drawString(120,600,str(quotationItem.currencyQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,590,'PESO APROX :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,580,'DOCUMENTO :')
            can.setFont('Helvetica',7)
            can.drawString(120,580,str(quotationItem.relatedDocumentQuotation))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,570,'OBRA :')

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,560,'OBSERVACIONES :')
            can.setFont('Helvetica',6)
            can.drawString(120,560,str(quotationItem.commentQuotation))

            #Linea de separacion con los datos del vendedor
            can.line(25,550,580,550)

            #Datos del vendedor
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,540,'VENDEDOR :')
            can.setFont('Helvetica',7)
            can.drawString(120,540,str(quotationItem.quotationsellerdata.dataUserQuotation[1]))
            can.setFont('Helvetica-Bold',7)
            can.drawString(25,530,'CELULAR :')
            can.setFont('Helvetica',7)
            can.drawString(120,530,str(quotationItem.quotationsellerdata.dataUserQuotation[3]))

            can.setFont('Helvetica-Bold',7)
            can.drawString(25,520,'EMAIL:')
            can.setFont('Helvetica',7)
            can.drawString(120,520,str(quotationItem.quotationsellerdata.asociatedUser.email))


            #Aqui se ponen las cabeceras

            can.setStrokeColorRGB(0,0,1)
            can.setFillColorRGB(0,0,0)
            #Campos en cabecera
            lista_x = [25,580]
            lista_y = [500,515]
            can.setFillColorRGB(0,0,1)
            can.rect(25,500,555,15,fill=1)

            #Valores iniciales
            lista_x = [25,50,120,310,360,410,460,530]
            lista_y = [500,515]

            #Ingreso de campo cantidad
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[0] + 5, lista_y[0] + 3,'Item.')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                can.drawRightString(lista_x[0] + 20,lista_y[0] + 3,str("1"))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de código de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[1] + 5, lista_y[0] + 3,'Código')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[1] + 5,lista_y[0] + 3,'Servicio')
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de descripcion de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[2] + 5, lista_y[0] + 3,'Descripción')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[2] + 5,lista_y[0] + 3,serviceInfo.dataServiceQuotation[1])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de unidad de medida de producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[3] + 5, lista_y[0] + 3,'Und')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                can.drawString(lista_x[3] + 5,lista_y[0] + 3,serviceInfo.dataServiceQuotation[2])
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            condicion_imprimir = quotationItem.showUnitPrice + quotationItem.showSellPrice + quotationItem.showDiscount
            if condicion_imprimir == 'ONOFFOFF':
                lista_x[4] = 360
            
            if condicion_imprimir == 'ONOFFON':
                lista_x[5] = 360

            if condicion_imprimir == 'OFFOFFON':
                lista_x[6] = 360
            
            if condicion_imprimir == 'ONONOFF':
                lista_x[4] = 360
                lista_x[5] = 420

            if condicion_imprimir == 'ONOFFON':
                lista_x[4] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == 'OFFONON':
                lista_x[5] = 360
                lista_x[6] = 420
            
            if condicion_imprimir == 'ONONON':
                lista_x[4] = 360
                lista_x[5] = 420
                lista_x[6] = 480


            if quotationItem.showUnitPrice == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de unidad de medida de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[4] - 5, lista_y[0] + 3,'V.U sin IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for serviceInfo in groupQuantity[indicatorGroup]:
                    if quotationItem.currencyQuotation == 'SOLES':
                        if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                        if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                    if quotationItem.currencyQuotation == 'DOLARES':
                        if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                            vu_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                        if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                    can.drawRightString(lista_x[4] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % vu_servicio)))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]

            if quotationItem.showSellPrice == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo del precio con IGV de producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[5] - 5, lista_y[0]+3,'P.U con IGV')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for serviceInfo in groupQuantity[indicatorGroup]:
                    if quotationItem.currencyQuotation == 'SOLES':
                        if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                        if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                    if quotationItem.currencyQuotation == 'DOLARES':
                        if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                            vu_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                        if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                            vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                    can.drawRightString(lista_x[5] + 20,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % (vu_servicio*Decimal(1.18)))))
                    lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            
            if quotationItem.showDiscount == 'ON':
                #Valores iniciales
                lista_y = [500,515]
                #Ingreso de campo de descuento del producto
                can.setFillColorRGB(1,1,1)
                can.setFont('Helvetica-Bold',7)
                can.drawString(lista_x[6], lista_y[0] + 3,'Dscto')
                can.setFont('Helvetica',7)
                can.setFillColorRGB(0,0,0)
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                for serviceInfo in groupQuantity[indicatorGroup]:
                    can.drawRightString(lista_x[6] + 20,lista_y[0] + 3,str(serviceInfo.dataServiceQuotation[5]) + ' %')
                    lista_y = [lista_y[0] - 15,lista_y[1] - 15]

            #Valores iniciales
            lista_y = [500,515]
            #Ingreso de campo de valor de venta del producto
            can.setFillColorRGB(1,1,1)
            can.setFont('Helvetica-Bold',7)
            can.drawString(lista_x[7] + 5, lista_y[0] + 3,'Valor Venta')
            can.setFont('Helvetica',7)
            can.setFillColorRGB(0,0,0)
            lista_y = [lista_y[0] - 16,lista_y[1] - 16]
            for serviceInfo in groupQuantity[indicatorGroup]:
                if quotationItem.currencyQuotation == 'SOLES':
                    if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                        vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                    if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                        vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                if quotationItem.currencyQuotation == 'DOLARES':
                    if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                        vu_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                    if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                        vu_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                #v_producto = round(v_producto,2)
                can.drawRightString(lista_x[7] + 45,lista_y[0] + 3,"{:,}".format(Decimal('%.2f' % Decimal(vu_servicio))))
                lista_y = [lista_y[0] - 16,lista_y[1] - 16]
                total_precio = Decimal(total_precio) + Decimal(vu_servicio)

            #Linea de separacion con los datos finales
            can.line(25,lista_y[1],580,lista_y[1])

            #Logo de las empresas
            can.drawImage('./static/salesMetalprotec/images/mega.png',20,60,width=90,height=60,mask='auto')
            can.drawImage('./static/salesMetalprotec/images/hodelpe.png',470,65,width=100,height=50,mask='auto')

            #Linea inicio de separacion
            can.line(25,60,580,60)

            #Impresion de los datos bancarios
            #Scotiabank
            can.setFont('Helvetica-Bold',8)
            can.drawString(25,50,'Banco Scotiabank')
            can.setFont('Helvetica',8)
            can.drawString(25,40,'Cta Cte Soles: 000 9496505')
            can.drawString(25,30,'Cta Cte Dolares: 000 5151261')

            #BCP
            can.setFont('Helvetica-Bold',8)
            can.drawString(230,50,'Banco de Crédito del Perú')
            can.setFont('Helvetica',8)
            can.drawString(230,40,'Cta Cte Soles: 310 9888337 0 02')
            can.drawString(230,30,'Cta Cte Dolares: 310 9865292 1 35')

            #BBVA
            can.setFont('Helvetica-Bold',8)
            can.drawString(420,50,'Banco Continental BBVA')
            can.setFont('Helvetica',8)
            can.drawString(420,40,'Cta Cte Soles: 0011 0250 0200615638 80')
            can.drawString(420,30,'Cta Cte Dolares: 0011 0250 0200653947 88')

            #Linea final de separacion
            can.line(25,25,580,25)

            indicatorGroup = indicatorGroup + 1
            if totalGroups > indicatorGroup:
                can.showPage()
        
        #Esta seccion solo va en la hoja final de los productos
        #Impresion de total venta
        can.drawRightString(480,lista_y[0]+4,'Total Venta Grabada')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % total_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de total IGV
        igv_precio = Decimal('%.2f' % total_precio)*Decimal(0.18)
        can.drawRightString(480,lista_y[0]+4,'Total IGV')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % igv_precio)))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Impresion de importe total
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'S/')
        else:
            can.drawRightString(490,lista_y[0]+4,'$')
        can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(precio_final))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion
        can.line(480,lista_y[1],580,lista_y[1])

        #Calculo de la proforma en dolares
        total_dolares = Decimal(0.0000)
        for serviceInfo in totalServicesQuotation:
            if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                v_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
            if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                v_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
            total_dolares = Decimal(total_dolares) + Decimal(v_servicio)
        final_dolares = Decimal('%.2f' % total_dolares)*Decimal(1.18)

        #Impresion de importe en otra moneda
        precio_final = Decimal('%.2f' % total_precio)*Decimal(1.18)
        can.drawRightString(480,lista_y[0]+4,'Importe Total de la Venta')
        if quotationItem.currencyQuotation == 'SOLES':
            can.drawRightString(490,lista_y[0]+4,'$')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % Decimal(final_dolares))))
        else:
            can.drawRightString(490,lista_y[0]+4,'S/')
            can.drawRightString(lista_x[7]+45,lista_y[0]+4,"{:,}".format(Decimal('%.2f' % (Decimal(precio_final)))))
        lista_y = [lista_y[0] - 15,lista_y[1] - 15]

        #Linea de separacion con los datos finales
        can.line(25,lista_y[1],580,lista_y[1])
        can.save()

    nombre_doc = str(quotationItem.codeQuotation) + '.pdf'
    response = HttpResponse(open('quotationMetalprotec.pdf','rb'),content_type='application/pdf')
    nombre = 'attachment; ' + 'filename=' + nombre_doc
    response['Content-Disposition'] = nombre
    return response

def sendGuideTeFacturo(request,idGuide):
    guideObject = guideSystem.objects.get(id=idGuide)
    infoGuide = getInfoGuide(guideObject)
    print(infoGuide)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlGuide = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/guia-remision'
        r = requests.post(urlGuide,headers=headersTeFacturo,json=infoGuide)
        print(r)
        print(r.content)
        if((r.status_code==200) or (r.status_code==201)):
            guideObject.stateGuide = 'ENVIADA'
        guideObject.save()
    else:
        guideObject.stateGuide = 'ENVIADA'
        guideObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:guidesMetalprotec'))

def verifyGuideTeFacturo(request,idGuide):
    guideObject = guideSystem.objects.get(id=idGuide)
    infoGuide = {
        'emisor':'20541628631',
        'numero':str(guideObject.nroGuide),
        'serie':str(guideObject.endpointGuide.serieGuia),
        'tipoComprobante':'09'
    }
    print(infoGuide)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlState = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarEstado'
        r = requests.put(urlState,headers=headersTeFacturo,json=infoGuide)
        guideObject.stateTeFacturo = r.json().get('estadoSunat').get('valor')
        guideObject.save()
    else:
        guideObject.stateTeFacturo = 'Aceptado con Obs.'
        guideObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:guidesMetalprotec'))

def verifyCreditNoteTeFacturo(request,idCreditNote):
    creditNoteObject = creditNoteSystem.objects.get(id=idCreditNote)
    if creditNoteObject.originCreditNote == 'INVOICE':
        serieCreditNote = creditNoteObject.endpointCreditNote.serieNotaBoleta
    else:
        serieCreditNote = creditNoteObject.endpointCreditNote.serieNotaFactura
    infoCreditNote = {
        'emisor':'20541628631',
        'numero':str(creditNoteObject.nroCreditNote),
        'serie':serieCreditNote,
        'tipoComprobante':'07'
    }
    print(infoCreditNote)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlState = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarEstado'
        r = requests.put(urlState,headers=headersTeFacturo,json=infoCreditNote)
        creditNoteObject.stateTeFacturo = r.json().get('estadoSunat').get('valor')
        creditNoteObject.save()
    else:
        creditNoteObject.stateTeFacturo = 'Aceptado con Obs.'
        creditNoteObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:creditNotesMetalprotec'))

def verifyBillTeFacturo(request,idBill):
    billObject = billSystem.objects.get(id=idBill)
    infoBill = {
        'emisor':'20541628631',
        'numero':str(billObject.nroBill),
        'serie':str(billObject.endpointBill.serieFactura),
        'tipoComprobante':'01'
    }
    print(infoBill)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlState = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarEstado'
        r = requests.put(urlState,headers=headersTeFacturo,json=infoBill)
        billObject.stateTeFacturo = r.json().get('estadoSunat').get('valor')
        billObject.save()
    else:
        billObject.stateTeFacturo = 'Aceptado con Obs.'
        billObject.save()

    if (billObject.stockBill != '2' and billObject.stockBill != '1') and billObject.typeItemsBill=='PRODUCTOS':
        billObject.stockBill = '1'
        billObject.save()
        if billObject.originBill == 'QUOTATION':
            asociatedQuotation = billObject.asociatedQuotation
            allProductsInfo = asociatedQuotation.quotationproductdata_set.all()
            try:
                for productInfo in allProductsInfo:
                    asociatedProduct = productInfo.asociatedProduct
                    storeObject = storeSystem.objects.get(nameStore=productInfo.dataProductQuotation[4])
                    stockEdit = storexproductSystem.objects.filter(asociatedProduct=asociatedProduct).get(asociatedStore=storeObject)
                    lastStock = stockEdit.quantityProduct
                    addStockQt = productInfo.dataProductQuotation[8]
                    stockEdit.quantityProduct = str(Decimal('%.2f' % Decimal(Decimal(stockEdit.quantityProduct) - Decimal(addStockQt))))
                    stockEdit.save()
                    newStock = stockEdit.quantityProduct
                    typeOutcoming = 'EGRESO-FACTURA'
                    dateOutcoming = datetime.datetime.today()
                    productCode = asociatedProduct.codeProduct
                    nameStore = storeObject.nameStore
                    quantityProduct = addStockQt
                    referenceOutcome = billObject.codeBill
                    asociatedUserData = request.user
                    asociatedProduct = asociatedProduct
                    asociatedStoreData = storeObject
                    endpointOutcoming = request.user.extendeduser.endpointUser
                    outcomingItemsRegisterInfo.objects.create(
                        typeOutcoming=typeOutcoming,
                        dateOutcoming=dateOutcoming,
                        productCode=productCode,
                        nameStore=nameStore,
                        quantityProduct=quantityProduct,
                        lastStock=lastStock,
                        newStock=newStock,
                        referenceOutcome=referenceOutcome,
                        asociatedUserData=asociatedUserData,
                        asociatedProduct=asociatedProduct,
                        asociatedBill=billObject,
                        asociatedStoreData=storeObject,
                        endpointOutcoming=endpointOutcoming
                    )
                billObject.stockBill = '2'
                billObject.save()
            except:
                print('Se ha fallado')
                billObject.stockBill = '1'
                billObject.save()

        else:
            asociatedQuotation = billObject.guidesystem_set.all()[0].asociatedQuotation
            allProductsInfo = asociatedQuotation.quotationproductdata_set.all()
            try:
                for productInfo in allProductsInfo:
                    asociatedProduct = productInfo.asociatedProduct
                    storeObject = storeSystem.objects.get(nameStore=productInfo.dataProductQuotation[4])
                    stockEdit = storexproductSystem.objects.filter(asociatedProduct=asociatedProduct).get(asociatedStore=storeObject)
                    lastStock = stockEdit.quantityProduct
                    addStockQt = productInfo.dataProductQuotation[8]
                    stockEdit.quantityProduct = str(Decimal('%.2f' % Decimal(Decimal(stockEdit.quantityProduct) - Decimal(addStockQt))))
                    stockEdit.save()
                    newStock = stockEdit.quantityProduct
                    typeOutcoming = 'EGRESO-FACTURA'
                    dateOutcoming = datetime.datetime.today()
                    productCode = asociatedProduct.codeProduct
                    nameStore = storeObject.nameStore
                    quantityProduct = addStockQt
                    referenceOutcome = billObject.codeBill
                    asociatedUserData = request.user
                    asociatedProduct = asociatedProduct
                    asociatedStoreData = storeObject
                    endpointOutcoming = request.user.extendeduser.endpointUser
                    outcomingItemsRegisterInfo.objects.create(
                        typeOutcoming=typeOutcoming,
                        dateOutcoming=dateOutcoming,
                        productCode=productCode,
                        nameStore=nameStore,
                        quantityProduct=quantityProduct,
                        lastStock=lastStock,
                        newStock=newStock,
                        referenceOutcome=referenceOutcome,
                        asociatedUserData=asociatedUserData,
                        asociatedProduct=asociatedProduct,
                        asociatedBill=billObject,
                        asociatedStoreData=storeObject,
                        endpointOutcoming=endpointOutcoming
                    )
                billObject.stockBill = '2'
                billObject.save()
            except:
                print('Se ha fallado')
                billObject.stockBill = '1'
                billObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:billsMetalprotec'))

def verifyInvoiceTeFacturo(request,idInvoice):
    invoiceObject = invoiceSystem.objects.get(id=idInvoice)
    infoInvoice = {
        'emisor':'20541628631',
        'numero':str(invoiceObject.nroInvoice),
        'serie':str(invoiceObject.endpointInvoice.serieBoleta),
        'tipoComprobante':'03'
    }
    print(infoInvoice)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlState = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarEstado'
        r = requests.put(urlState,headers=headersTeFacturo,json=infoInvoice)
        print(r.json().get('estadoSunat').get('valor'))
        invoiceObject.stateTeFacturo = r.json().get('estadoSunat').get('valor')
        invoiceObject.save()
    else:
        invoiceObject.stateTeFacturo = 'Aceptado con Obs.'
        invoiceObject.save()
    if (invoiceObject.stockInvoice != '2' and invoiceObject.stockInvoice != '1') and invoiceObject.typeItemsInvoice=='PRODUCTOS':
        invoiceObject.stockInvoice = '1'
        invoiceObject.save()
        if invoiceObject.originInvoice == 'QUOTATION':
            asociatedQuotation = invoiceObject.asociatedQuotation
            allProductsInfo = asociatedQuotation.quotationproductdata_set.all()
            try:
                for productInfo in allProductsInfo:
                    asociatedProduct = productInfo.asociatedProduct
                    storeObject = storeSystem.objects.get(nameStore=productInfo.dataProductQuotation[4])
                    stockEdit = storexproductSystem.objects.filter(asociatedProduct=asociatedProduct).get(asociatedStore=storeObject)
                    lastStock = stockEdit.quantityProduct
                    addStockQt = productInfo.dataProductQuotation[8]
                    stockEdit.quantityProduct = str(Decimal('%.2f' % Decimal(Decimal(stockEdit.quantityProduct) - Decimal(addStockQt))))
                    stockEdit.save()
                    newStock = stockEdit.quantityProduct
                    typeOutcoming = 'EGRESO-INVOICE'
                    dateOutcoming = datetime.datetime.today()
                    productCode = asociatedProduct.codeProduct
                    nameStore = storeObject.nameStore
                    quantityProduct = addStockQt
                    referenceOutcome = invoiceObject.codeInvoice
                    asociatedUserData = request.user
                    asociatedProduct = asociatedProduct
                    asociatedStoreData = storeObject
                    endpointOutcoming = request.user.extendeduser.endpointUser
                    outcomingItemsRegisterInfo.objects.create(
                        typeOutcoming=typeOutcoming,
                        dateOutcoming=dateOutcoming,
                        productCode=productCode,
                        nameStore=nameStore,
                        quantityProduct=quantityProduct,
                        lastStock=lastStock,
                        newStock=newStock,
                        referenceOutcome=referenceOutcome,
                        asociatedUserData=asociatedUserData,
                        asociatedProduct=asociatedProduct,
                        asociatedInvoice=invoiceObject,
                        asociatedStoreData=storeObject,
                        endpointOutcoming=endpointOutcoming
                    )
                invoiceObject.stockInvoice = '2'
                invoiceObject.save()
            except:
                print('Se ha fallado')
                invoiceObject.stockInvoice = '1'
                invoiceObject.save()

        else:
            asociatedQuotation = invoiceObject.guidesystem_set.all()[0].asociatedQuotation
            allProductsInfo = asociatedQuotation.quotationproductdata_set.all()
            try:
                for productInfo in allProductsInfo:
                    asociatedProduct = productInfo.asociatedProduct
                    storeObject = storeSystem.objects.get(nameStore=productInfo.dataProductQuotation[4])
                    stockEdit = storexproductSystem.objects.filter(asociatedProduct=asociatedProduct).get(asociatedStore=storeObject)
                    lastStock = stockEdit.quantityProduct
                    addStockQt = productInfo.dataProductQuotation[8]
                    stockEdit.quantityProduct = str(Decimal('%.2f' % Decimal(Decimal(stockEdit.quantityProduct) - Decimal(addStockQt))))
                    stockEdit.save()
                    newStock = stockEdit.quantityProduct
                    typeOutcoming = 'EGRESO-INVOICE'
                    dateOutcoming = datetime.datetime.today()
                    productCode = asociatedProduct.codeProduct
                    nameStore = storeObject.nameStore
                    quantityProduct = addStockQt
                    referenceOutcome = invoiceObject.codeInvoice
                    asociatedUserData = request.user
                    asociatedProduct = asociatedProduct
                    asociatedStoreData = storeObject
                    endpointOutcoming = request.user.extendeduser.endpointUser
                    outcomingItemsRegisterInfo.objects.create(
                        typeOutcoming=typeOutcoming,
                        dateOutcoming=dateOutcoming,
                        productCode=productCode,
                        nameStore=nameStore,
                        quantityProduct=quantityProduct,
                        lastStock=lastStock,
                        newStock=newStock,
                        referenceOutcome=referenceOutcome,
                        asociatedUserData=asociatedUserData,
                        asociatedProduct=asociatedProduct,
                        asociatedInvoice=invoiceObject,
                        asociatedStoreData=storeObject,
                        endpointOutcoming=endpointOutcoming
                    )
                invoiceObject.stockInvoice = '2'
                invoiceObject.save()
            except:
                print('Se ha fallado')
                invoiceObject.stockInvoice = '1'
                invoiceObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:invoicesMetalprotec'))

def downloadGuideTeFacturo(request,idGuide):
    guideObject = guideSystem.objects.get(id=idGuide)
    infoGuide = {
        "emisor":"20541628631",
        "numero":int(guideObject.nroGuide),
        "serie":guideObject.endpointGuide.serieGuia,
        "tipoComprobante":"09"
    }
    print(infoGuide)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlDownload = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarPdf'
        r = requests.put(urlDownload,headers=headersTeFacturo,json=infoGuide)
        convert_b64 = r.content
        info_decoded = b64decode(convert_b64,validate=True)

        if info_decoded[0:4] != b'%PDF':
            print("Hay un error en el pdf")
            raise ValueError('Missing the PDF file signature')
        print('PDF LISTO PARA ENVIAR')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; ' + f'filename={guideObject.codeGuide}.pdf'
        response.write(info_decoded)
        return response
    else:
        return HttpResponseRedirect(reverse('salesMetalprotec:guidesMetalprotec'))

def downloadCreditNoteTeFacturo(request,idCreditNote):
    creditNoteObject = creditNoteSystem.objects.get(id=idCreditNote)
    if creditNoteObject.originCreditNote == 'INVOICE':
        serieCreditNote = creditNoteObject.endpointCreditNote.serieNotaBoleta
    else:
        serieCreditNote = creditNoteObject.endpointCreditNote.serieNotaFactura
    infoCreditNote = {
        "emisor":"20541628631",
        "numero":int(creditNoteObject.nroCreditNote),
        "serie":serieCreditNote,
        "tipoComprobante":"07"
    }
    print(infoCreditNote)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlDownload = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarPdf'
        r = requests.put(urlDownload,headers=headersTeFacturo,json=infoCreditNote)
        convert_b64 = r.content
        info_decoded = b64decode(convert_b64,validate=True)

        if info_decoded[0:4] != b'%PDF':
            print("Hay un error en el pdf")
            raise ValueError('Missing the PDF file signature')
        print('PDF LISTO PARA ENVIAR')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; ' + f'filename={creditNoteObject.codeCreditNote}.pdf'
        response.write(info_decoded)
        return response
    else:
        return HttpResponseRedirect(reverse('salesMetalprotec:creditNotesMetalprotec'))
    
def downloadBillTeFacturo(request,idBill):
    billObject = billSystem.objects.get(id=idBill)
    infoBill = {
        "emisor":"20541628631",
        "numero":int(billObject.nroBill),
        "serie":billObject.endpointBill.serieFactura,
        "tipoComprobante":"01"
    }
    print(infoBill)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlDownload = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarPdf'
        r = requests.put(urlDownload,headers=headersTeFacturo,json=infoBill)
        convert_b64 = r.content
        info_decoded = b64decode(convert_b64,validate=True)

        if info_decoded[0:4] != b'%PDF':
            print("Hay un error en el pdf")
            raise ValueError('Missing the PDF file signature')
        print('PDF LISTO APRA ENVIAR')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; ' + f'filename={billObject.codeBill}.pdf'
        response.write(info_decoded)
        return response
    else:
        return HttpResponseRedirect(reverse('salesMetalprotec:billsMetalprotec'))

def downloadInvoiceTeFacturo(request,idInvoice):
    invoiceObject = invoiceSystem.objects.get(id=idInvoice)
    infoInvoice = {
        "emisor":"20541628631",
        "numero":int(invoiceObject.nroInvoice),
        "serie":invoiceObject.endpointInvoice.serieBoleta,
        "tipoComprobante":"03"
    }
    print(infoInvoice)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlDownload = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/consultarPdf'
        r = requests.put(urlDownload,headers=headersTeFacturo,json=infoInvoice)
        convert_b64 = r.content
        info_decoded = b64decode(convert_b64,validate=True)

        if info_decoded[0:4] != b'%PDF':
            print("Hay un error en el pdf")
            raise ValueError('Missing the PDF file signature')
        print('PDF LISTO APRA ENVIAR')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; ' + f'filename={invoiceObject.codeInvoice}.pdf'
        response.write(info_decoded)
        return response
    else:
        return HttpResponseRedirect(reverse('salesMetalprotec:invoicesMetalprotec'))

def getInfoGuide(guideObject):
    totalProductsInfo = guideObject.asociatedQuotation.quotationproductdata_set.all()

    #Weight of the products
    totalWeight = 0.00
    for productInfo in totalProductsInfo:
        totalWeight = totalWeight + round(float(productInfo.dataProductQuotation[10])*float(productInfo.dataProductQuotation[8]),2)
    totalWeight = totalWeight + float(guideObject.extraWeight)
    totalWeight = str(int(totalWeight))

    #Getting productsGuide
    totalProductsGuide = []
    for productInfo in totalProductsInfo:
        if productInfo.asociatedProduct.kitProduct == 'ON':
            for infoProductKit in productInfo.asociatedProduct.kitInfo:
                productInfoKit = productSystem.objects.get(id=infoProductKit[0])
                quantityKit = int(float(infoProductKit[1]))*int(float(productInfo.dataProductQuotation[8]))
                totalProductsGuide.append([productInfoKit.nameProduct,productInfoKit.codeProduct,quantityKit])
        else:
            totalProductsGuide.append([productInfo.dataProductQuotation[1],productInfo.dataProductQuotation[2],int(float(productInfo.dataProductQuotation[8]))])

    productsTeFacturo = []
    i = 1
    for productoItem in totalProductsGuide:
        infoProductsItems = {
            "numeroOrden":i,
            "cantidad":productoItem[2],
            "codigoProducto":productoItem[1],
            "descripcion":productoItem[0],
            "unidadMedida": "UNIDAD_BIENES"
        }
        productsTeFacturo.append(infoProductsItems)
        i = i + 1

    if guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[3] == 'PERSONA':
        destinatario = {
            "nombreLegal":str(guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[1]),
            "numeroDocumentoIdentidad":str(guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[2]),
            "tipoDocumentoIdentidad": "DOC_NACIONAL_DE_IDENTIDAD",
            "correo":str(guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[4])
        }
    else:
        destinatario = {
            "nombreLegal":str(guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[1]),
            "numeroDocumentoIdentidad":str(guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[2]),
            "tipoDocumentoIdentidad": "RUC",
            "correo":str(guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[4])
        }

    if guideObject.modeTransportation == 'PRIVADO':
        guia_vehiculos = [
            {
                "placa":guideObject.vehiclePlate,
                "liscenciaConducir":guideObject.licenceDriver,
                "conductor":
                {
                    "nombreLegal":guideObject.nameDriver,
                    "numeroDocumentoIdentidad":guideObject.dniDriver,
                    "tipoDocumentoIdentidad":"DOC_NACIONAL_DE_IDENTIDAD",
                }
            }
        ]

        param_data = {
            "close2u": 
            {
                "tipoIntegracion": "OFFLINE",
                "tipoPlantilla": "01",
                "tipoRegistro": "PRECIOS_SIN_IGV"
            },
            "datosDocumento":
            {
                "fechaEmision": guideObject.dateGuide.strftime("%Y-%m-%d"),
                "glosa": guideObject.commentGuide,
                "numero": int(guideObject.nroGuide),
                "serie": guideObject.endpointGuide.serieGuia
            },
            "remitente":
            {
                "nombreLegal": "METALPROTEC",
                "numeroDocumentoIdentidad": "20541628631",
                "tipoDocumentoIdentidad": "RUC"
            },
            "destinatario":destinatario,
            "datosEnvio":
            {
                "motivoTraslado": guideObject.purposeTransportation,
                "descripcionTraslado": "",
                "transbordoProgramado": "False",
                "pesoBruto": totalWeight,
                "unidadMedida": "KILOS",
                "numeroPallet": "0",
                "modalidadTraslado": guideObject.modeTransportation,
                "fechaTraslado": guideObject.dateGivenGoods.strftime("%Y-%m-%d"),
                "fechaEntrega": guideObject.dateGivenGoods.strftime("%Y-%m-%d"),
                "puntoLlegada":
                {
                    "departamento": "",
                    "direccion": guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[8],
                    "distrito": "",
                    "pais": "PERU",
                    "provincia": "",
                    "ubigeo": guideObject.ubigeoClient,
                    "urbanizacion": ""
                },
                "puntoPartida":
                {
                    "departamento": guideObject.deparmentDeparture,
                    "direccion": guideObject.addressDeparture,
                    "distrito": guideObject.districtDeparture,
                    "pais": "PERU",
                    "provincia": guideObject.provinceDeparture,
                    "ubigeo": guideObject.ubigeoDeparture,
                    "urbanizacion": ""
                },
                "numeroContenedor": ""
            },
            "vehiculos": guia_vehiculos,
            "detalleGuia": productsTeFacturo
        }
    
    if guideObject.modeTransportation == 'PUBLICO':
        guia_transportista = {
            "nombreLegal": guideObject.razonSocialTranporter,
            "numeroDocumentoIdentidad": guideObject.rucTransporter,
            "tipoDocumentoIdentidad": "RUC"
        }
        param_data = {
            "close2u": 
            {
                "tipoIntegracion": "OFFLINE",
                "tipoPlantilla": "01",
                "tipoRegistro": "PRECIOS_SIN_IGV"
            },
            "datosDocumento":
            {
                "fechaEmision": guideObject.dateGuide.strftime("%Y-%m-%d"),
                "glosa": guideObject.commentGuide,
                "numero": int(guideObject.nroGuide),
                "serie": guideObject.endpointGuide.serieGuia
            },
            "remitente":
            {
                "nombreLegal": "METALPROTEC",
                "numeroDocumentoIdentidad": "20541628631",
                "tipoDocumentoIdentidad": "RUC"
            },
            "destinatario":destinatario,
            "datosEnvio":
            {
                "motivoTraslado": guideObject.purposeTransportation,
                "descripcionTraslado": "",
                "transbordoProgramado": "False",
                "pesoBruto": totalWeight,
                "unidadMedida": "KILOS",
                "numeroPallet": "0",
                "modalidadTraslado": guideObject.modeTransportation,
                "fechaTraslado": guideObject.dateGivenGoods.strftime("%Y-%m-%d"),
                "fechaEntrega": guideObject.dateGivenGoods.strftime("%Y-%m-%d"),
                "puntoLlegada":
                {
                    "departamento": "",
                    "direccion": guideObject.asociatedQuotation.quotationclientdata.dataClientQuotation[8],
                    "distrito": "",
                    "pais": "PERU",
                    "provincia": "",
                    "ubigeo": guideObject.ubigeoClient,
                    "urbanizacion": ""
                },
                "puntoPartida":
                {
                    "departamento": guideObject.deparmentDeparture,
                    "direccion": guideObject.addressDeparture,
                    "distrito": guideObject.districtDeparture,
                    "pais": "PERU",
                    "provincia": guideObject.provinceDeparture,
                    "ubigeo": guideObject.ubigeoDeparture,
                    "urbanizacion": ""
                },
                "numeroContenedor": ""
            },
            "transportista": guia_transportista,
            "detalleGuia": productsTeFacturo
        }
    return param_data

def sendBillTeFacturo(request,idBill):
    billObject = billSystem.objects.get(id=idBill)
    infoBill = getInfoBill(billObject)
    print(infoBill)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlBill = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/factura'
        r = requests.put(urlBill,headers=headersTeFacturo,json=infoBill)
        print(r)
        print(r.content)
        if((r.status_code==200) or (r.status_code==201)):
            billObject.stateBill = 'ENVIADA'
            try:
                quotationItem = None
                if billObject.originBill == 'GUIDE':
                    for guideItem in billObject.guidesystem_set.all():
                        quotationItem = guideItem.asociatedQuotation
                        quotationItem.currencyQuotation = billObject.currencyBill
                        quotationItem.save()
                else:
                    quotationItem = billObject.asociatedQuotation
                    quotationItem.currencyQuotation = billObject.currencyBill
                    quotationItem.save()
            except:
                pass
        billObject.save()
    else:
        billObject.stateBill = 'ENVIADA'
        billObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:billsMetalprotec'))

def sendCreditNoteTeFacturo(request,idCreditNote):
    creditNoteObject = creditNoteSystem.objects.get(id=idCreditNote)
    infoCreditNote = getInfoCreditNote(creditNoteObject)
    print(infoCreditNote)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlCreditNote = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/nota-credito'
        r = requests.put(urlCreditNote,headers=headersTeFacturo,json=infoCreditNote)
        print(r)
        print(r.content)
        if((r.status_code==200) or (r.status_code==201)):
            creditNoteObject.stateCreditNote = 'ENVIADA'
        creditNoteObject.save()
    else:
        creditNoteObject.stateCreditNote = 'ENVIADA'
        creditNoteObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:creditNotesMetalprotec'))

def sendInvoiceTeFacturo(request,idInvoice):
    invoiceObject = invoiceSystem.objects.get(id=idInvoice)
    infoInvoice = getInfoInvoice(invoiceObject)
    print(infoInvoice)
    if env('ENV_PROJECT') == 'ProdMetalprotec':
        headersTeFacturo = {"X-Auth-Token":token_metalprotec,"Content-Type":"application/json"}
        urlInvoice = 'https://invoice2u.pe/apiemisor/invoice2u/integracion/boleta'
        r = requests.put(urlInvoice,headers=headersTeFacturo,json=infoInvoice)
        print(r)
        print(r.content)
        if((r.status_code==200) or (r.status_code==201)):
            invoiceObject.stateInvoice = 'ENVIADA'
            try:
                quotationItem = None
                if invoiceObject.originInvoice == 'GUIDE':
                    for guideItem in invoiceObject.guidesystem_set.all():
                        quotationItem = guideItem.asociatedQuotation
                        quotationItem.currencyQuotation = invoiceObject.currencyInvoice
                        quotationItem.save()
                else:
                    quotationItem = invoiceObject.asociatedQuotation
                    quotationItem.currencyQuotation = invoiceObject.currencyInvoice
                    quotationItem.save()
            except:
                pass
        invoiceObject.save()
    else:
        invoiceObject.stateInvoice = 'ENVIADA'
        invoiceObject.save()
    return HttpResponseRedirect(reverse('salesMetalprotec:invoicesMetalprotec'))

def getInfoBill(billObject):
    if billObject.typeItemsBill == 'SERVICIOS':
        refSuperior = {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"02",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        }
    else:
        refSuperior = {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"01",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        }

    if billObject.currencyBill == 'DOLARES':
        currencyBill = 'USD'
    else:
        currencyBill = 'PEN'

    if billObject.originBill == 'GUIDE':
        emailClient = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[4]
        addressClient = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[7]
        identificationClient = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[1]
        documentClient = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[2]
        condicionPago = billObject.guidesystem_set.all()[0].asociatedQuotation.paymentQuotation
        sellerInfo = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationsellerdata.dataUserQuotation[1]
        guideJson = []
        for guideInfo in billObject.guidesystem_set.all():
            newGuide = {
                "tipoDocumento":"GUIAEMISIONREMITENTE",
                "numero":guideInfo.nroGuide,
                "serie":billObject.endpointBill.serieGuia,
            }
            guideJson.append(newGuide)
    else:
        emailClient = billObject.asociatedQuotation.quotationclientdata.dataClientQuotation[4]
        addressClient = billObject.asociatedQuotation.quotationclientdata.dataClientQuotation[7]
        identificationClient = billObject.asociatedQuotation.quotationclientdata.dataClientQuotation[1]
        documentClient = billObject.asociatedQuotation.quotationclientdata.dataClientQuotation[2]
        condicionPago = billObject.asociatedQuotation.paymentQuotation
        sellerInfo = billObject.asociatedQuotation.quotationsellerdata.dataUserQuotation[1]
        guideJson = None

    itemsBill = []
    if billObject.typeItemsBill == 'PRODUCTOS':
        totalProductsItems = []
        if billObject.originBill == 'GUIDE':
            for guideInfo in billObject.guidesystem_set.all():
                for productInfo in guideInfo.asociatedQuotation.quotationproductdata_set.all():
                    if productInfo.asociatedProduct is not None:
                        if productInfo.asociatedProduct.kitProduct == 'ON':
                            productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                            totalProductsItems.append([
                                productoDatos.id,
                                productoDatos.nameProduct,
                                productoDatos.codeProduct,
                                productoDatos.measureUnit,
                                productInfo.dataProductQuotation[4],
                                productInfo.dataProductQuotation[5],
                                productInfo.dataProductQuotation[6],
                                productInfo.dataProductQuotation[7],
                                productInfo.dataProductQuotation[8],
                                productInfo.dataProductQuotation[9],
                            ])
                        else:
                            totalProductsItems.append(productInfo.dataProductQuotation)
                    else:
                        totalProductsItems.append(productInfo.dataProductQuotation)
        else:
            for productInfo in billObject.asociatedQuotation.quotationproductdata_set.all():
                if productInfo.asociatedProduct is not None:
                    if productInfo.asociatedProduct.kitProduct == 'ON':
                        productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                        totalProductsItems.append([
                            productoDatos.id,
                            productoDatos.nameProduct,
                            productoDatos.codeProduct,
                            productoDatos.measureUnit,
                            productInfo.dataProductQuotation[4],
                            productInfo.dataProductQuotation[5],
                            productInfo.dataProductQuotation[6],
                            productInfo.dataProductQuotation[7],
                            productInfo.dataProductQuotation[8],
                            productInfo.dataProductQuotation[9],
                        ])
                    else:
                        totalProductsItems.append(productInfo.dataProductQuotation)
                else:
                    totalProductsItems.append(productInfo.dataProductQuotation)
        
        finalPrice = 0
        if currencyBill == 'PEN':
            i = 1
            for productoItem in totalProductsItems:
                if productoItem[5] == 'SOLES':
                    precioProducto = Decimal(productoItem[6])*Decimal(Decimal(1.00) - (Decimal(productoItem[7])/100))
                    precioProducto = float('%.2f' % precioProducto)
                else:
                    precioProducto = Decimal(productoItem[6])*Decimal(billObject.erSel)*Decimal(Decimal(1.00) - Decimal(productoItem[7])/100)
                    precioProducto = float('%.2f' % precioProducto)
                if productoItem[9] == '1':
                    infoProductoItem = {
                        "codigoProducto":productoItem[2],
                        "codigoProductoSunat":"",
                        "descripcion":productoItem[1],
                        "tipoAfectacion":"EXONERADO_TRANSFERENCIA_GRATUITA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(precioProducto[8]))),
                        "valorReferencialUnitarioItem":precioProducto,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice
                else:
                    infoProductoItem = {
                        "codigoProducto":productoItem[2],
                        "codigoProductoSunat":"",
                        "descripcion":productoItem[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(productoItem[8]))),
                        "valorVentaUnitarioItem":precioProducto,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice + precioProducto*round(float(productoItem[8]),2)
                itemsBill.append(infoProductoItem)
                i = i +1
        else:
            i = 1
            for productoItem in totalProductsItems:
                if productoItem[5] == 'SOLES':
                    precioProducto = (Decimal(productoItem[6])/Decimal(billObject.erSel))*Decimal(Decimal(1.00) - (Decimal(productoItem[7])/100))
                    precioProducto = float('%.2f' % precioProducto)
                else:
                    precioProducto = Decimal(productoItem[6])*Decimal(Decimal(1.00) - Decimal(productoItem[7])/100)
                    precioProducto = float('%.2f' % precioProducto)
                if productoItem[9] == '1':
                    infoProductoItem = {
                        "codigoProducto":productoItem[2],
                        "codigoProductoSunat":"",
                        "descripcion":productoItem[1],
                        "tipoAfectacion":"EXONERADO_TRANSFERENCIA_GRATUITA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(precioProducto[8]))),
                        "valorReferencialUnitarioItem":precioProducto,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice
                else:
                    infoProductoItem = {
                        "codigoProducto":productoItem[2],
                        "codigoProductoSunat":"",
                        "descripcion":productoItem[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(productoItem[8]))),
                        "valorVentaUnitarioItem":precioProducto,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice + precioProducto*round(float(productoItem[8]),2)
                itemsBill.append(infoProductoItem)
                i = i +1
    else:
        totalServicesItems = []
        for serviceInfo in billObject.asociatedQuotation.quotationservicedata_set.all():
            totalServicesItems.append(serviceInfo.dataServiceQuotation)
        finalPrice = 0
        if currencyBill == 'PEN':
            i = 1
            for serviceItem in totalServicesItems:
                if serviceItem[3] == 'SOLES':
                    precioServicio = Decimal(serviceItem[4])*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                    precioServicio = float('%.2f' % precioServicio)
                if serviceItem[3] == 'DOLARES':
                    precioServicio = Decimal(serviceItem[4])*Decimal(billObject.erSel)*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                    precioServicio = float('%.2f' % precioServicio)
                infoServiceItem = {
                    "codigoProducto":serviceItem[0],
                    "codigoProductoSunat":"",
                    "descripcion":serviceItem[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_BIENES",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precioServicio,
                    "descuento":{
                        "monto":'0',
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                finalPrice = finalPrice + precioServicio
                itemsBill.append(infoServiceItem)
                i = i + 1
        else:
            i = 1
            for serviceItem in totalServicesItems:
                if serviceItem[3] == 'DOLARES':
                    precioServicio = Decimal(serviceItem[4])*Decimal(Decimal(1.00) - Decimal(serviceItem[5])/100)
                    precioServicio = float('%.2f' % precioServicio)
                if serviceItem[3] == 'SOLES':
                    precioServicio = (Decimal(serviceItem[4])/Decimal(billObject.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                    precioServicio = float('%.2f' % precioServicio)
                infoServiceItem = {
                    "codigoProducto":serviceItem[0],
                    "codigoProductoSunat":"",
                    "descripcion":serviceItem[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_BIENES",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precioServicio,
                    "descuento":{
                        "monto":'0',
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                finalPrice = finalPrice + precioServicio
                itemsBill.append(infoServiceItem)
                i = i + 1
    
    quotesData = []
    if condicionPago == 'CONTADO':
        quotesData = None
    else:
        totalQuotes = len(billObject.dateQuotesBill)
        finalPrice = float('%.2f' % finalPrice)
        monto = '%.2f' % ((finalPrice*1.18)/int(totalQuotes))
        i=0
        while i < totalQuotes:
            quoteInfo = {
                "numero":'00' + str(i+1),
                "fecha":billObject.dateQuotesBill[i],
                "monto":str(monto),
                "moneda":currencyBill
            }
            i = i+1
            quotesData.append(quoteInfo)
    
    param_data = {
        "close2u":refSuperior, 
        "datosDocumento":
        {
            "serie":billObject.endpointBill.serieFactura,
            "numero":billObject.nroBill,
            "moneda":currencyBill,
            "fechaEmision":billObject.dateBill.strftime("%Y-%m-%d"),
            "horaEmision":None,
            "fechaVencimiento": None,
            "formaPago":condicionPago,
            "medioPago": "DEPOSITO_CUENTA",
            "condicionPago": None,
            "ordencompra":billObject.relatedDocumentBill,
            "puntoEmisor":None,
            "glosa":str(billObject.commentBill),
        },
        "detalleDocumento":itemsBill,
        "emisor":
        {
            "correo":"info@metalprotec.com",
            "nombreComercial": None,
            "nombreLegal": "METALPROTEC",
            "numeroDocumentoIdentidad": "20541628631",
            "tipoDocumentoIdentidad": "RUC",
            "domicilioFiscal":
            {
                "pais":"PERU",
                "departamento":"ANCASH",
                "provincia":"SANTA",
                "distrito":"NUEVO CHIMBOTE",
                "direccon":"Mza. J4 Lote. 39",
                "ubigeo":"02712"
            }
        },
        "informacionAdicional":
        {
            "tipoOperacion":"VENTA_INTERNA",
            "coVendedor":None,
            "vendedor":sellerInfo
        },
        "receptor":
        {
            "correo":emailClient,
            "correoCopia":None,
            "domicilioFiscal":
            {
                "direccion":addressClient,
                "pais":"PERU",
            },
            "nombreComercial":None,
            "nombreLegal":identificationClient,
            "numeroDocumentoIdentidad":documentClient,
            "tipoDocumentoIdentidad":"RUC"
        },
        'referencias':{
            "documentoReferenciaList":guideJson,
        },
        'cuotas':quotesData
    }
    return param_data

def getInfoInvoice(invoiceObject):
    if invoiceObject.typeItemsInvoice == 'SERVICIOS':
        refSuperior = {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"02",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        }
    else:
        refSuperior = {
            "tipoIntegracion":"OFFLINE",
            "tipoPlantilla":"01",
            "tipoRegistro":"PRECIOS_SIN_IGV"
        }

    if invoiceObject.currencyInvoice == 'DOLARES':
        currencyInvoice = 'USD'
    else:
        currencyInvoice = 'PEN'

    if invoiceObject.originInvoice == 'GUIDE':
        emailClient = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[4]
        addressClient = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[7]
        identificationClient = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[1]
        documentClient = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[2]
        condicionPago = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.paymentQuotation
        sellerInfo = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationsellerdata.dataUserQuotation[1]
        guideJson = []
        for guideInfo in invoiceObject.guidesystem_set.all():
            newGuide = {
                "tipoDocumento":"GUIAEMISIONREMITENTE",
                "numero":guideInfo.nroGuide,
                "serie":invoiceObject.endpointInvoice.serieGuia,
            }
            guideJson.append(newGuide)
    else:
        emailClient = invoiceObject.asociatedQuotation.quotationclientdata.dataClientQuotation[4]
        addressClient = invoiceObject.asociatedQuotation.quotationclientdata.dataClientQuotation[7]
        identificationClient = invoiceObject.asociatedQuotation.quotationclientdata.dataClientQuotation[1]
        documentClient = invoiceObject.asociatedQuotation.quotationclientdata.dataClientQuotation[2]
        condicionPago = invoiceObject.asociatedQuotation.paymentQuotation
        sellerInfo = invoiceObject.asociatedQuotation.quotationsellerdata.dataUserQuotation[1]
        guideJson = None

    itemsInvoice = []
    if invoiceObject.typeItemsInvoice == 'PRODUCTOS':
        totalProductsItems = []
        if invoiceObject.originInvoice == 'GUIDE':
            for guideInfo in invoiceObject.guidesystem_set.all():
                for productInfo in guideInfo.asociatedQuotation.quotationproductdata_set.all():
                    if productInfo.asociatedProduct is not None:
                        if productInfo.asociatedProduct.kitProduct == 'ON':
                            productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                            totalProductsItems.append([
                                productoDatos.id,
                                productoDatos.nameProduct,
                                productoDatos.codeProduct,
                                productoDatos.measureUnit,
                                productInfo.dataProductQuotation[4],
                                productInfo.dataProductQuotation[5],
                                productInfo.dataProductQuotation[6],
                                productInfo.dataProductQuotation[7],
                                productInfo.dataProductQuotation[8],
                                productInfo.dataProductQuotation[9],
                            ])
                        else:
                            totalProductsItems.append(productInfo.dataProductQuotation)
                    else:
                        totalProductsItems.append(productInfo.dataProductQuotation)
        else:
            for productInfo in invoiceObject.asociatedQuotation.quotationproductdata_set.all():
                if productInfo.asociatedProduct is not None:
                    if productInfo.asociatedProduct.kitProduct == 'ON':
                        productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                        totalProductsItems.append([
                            productoDatos.id,
                            productoDatos.nameProduct,
                            productoDatos.codeProduct,
                            productoDatos.measureUnit,
                            productInfo.dataProductQuotation[4],
                            productInfo.dataProductQuotation[5],
                            productInfo.dataProductQuotation[6],
                            productInfo.dataProductQuotation[7],
                            productInfo.dataProductQuotation[8],
                            productInfo.dataProductQuotation[9],
                        ])
                    else:
                        totalProductsItems.append(productInfo.dataProductQuotation)
                else:
                    totalProductsItems.append(productInfo.dataProductQuotation)
        
        finalPrice = 0
        if currencyInvoice == 'PEN':
            i = 1
            for productoItem in totalProductsItems:
                if productoItem[5] == 'SOLES':
                    precioProducto = Decimal(productoItem[6])*Decimal(Decimal(1.00) - (Decimal(productoItem[7])/100))
                    precioProducto = float('%.2f' % precioProducto)
                else:
                    precioProducto = Decimal(productoItem[6])*Decimal(invoiceObject.erSel)*Decimal(Decimal(1.00) - Decimal(productoItem[7])/100)
                    precioProducto = float('%.2f' % precioProducto)
                if productoItem[9] == '1':
                    infoProductoItem = {
                        "codigoProducto":productoItem[2],
                        "codigoProductoSunat":"",
                        "descripcion":productoItem[1],
                        "tipoAfectacion":"EXONERADO_TRANSFERENCIA_GRATUITA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(precioProducto[8]))),
                        "valorReferencialUnitarioItem":precioProducto,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice
                else:
                    infoProductoItem = {
                        "codigoProducto":productoItem[2],
                        "codigoProductoSunat":"",
                        "descripcion":productoItem[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(productoItem[8]))),
                        "valorVentaUnitarioItem":precioProducto,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice + precioProducto*round(float(productoItem[8]),2)
                itemsInvoice.append(infoProductoItem)
                i = i +1
        else:
            i = 1
            for productoItem in totalProductsItems:
                if productoItem[5] == 'SOLES':
                    precioProducto = (Decimal(productoItem[6])/Decimal(invoiceObject.erSel))*Decimal(Decimal(1.00) - (Decimal(productoItem[7])/100))
                    precioProducto = float('%.2f' % precioProducto)
                else:
                    precioProducto = Decimal(productoItem[6])*Decimal(Decimal(1.00) - Decimal(productoItem[7])/100)
                    precioProducto = float('%.2f' % precioProducto)
                if productoItem[9] == '1':
                    infoProductoItem = {
                        "codigoProducto":productoItem[2],
                        "codigoProductoSunat":"",
                        "descripcion":productoItem[1],
                        "tipoAfectacion":"EXONERADO_TRANSFERENCIA_GRATUITA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(precioProducto[8]))),
                        "valorReferencialUnitarioItem":precioProducto,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice
                else:
                    infoProductoItem = {
                        "codigoProducto":productoItem[2],
                        "codigoProductoSunat":"",
                        "descripcion":productoItem[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":str(int(float(productoItem[8]))),
                        "valorVentaUnitarioItem":precioProducto,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice + precioProducto*round(float(productoItem[8]),2)
                itemsInvoice.append(infoProductoItem)
                i = i +1
    else:
        totalServicesItems = []
        for serviceInfo in invoiceObject.asociatedQuotation.quotationservicedata_set.all():
            totalServicesItems.append(serviceInfo.dataServiceQuotation)
        finalPrice = 0
        if currencyInvoice == 'PEN':
            i = 1
            for serviceItem in totalServicesItems:
                if serviceItem[3] == 'SOLES':
                    precioServicio = Decimal(serviceItem[4])*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                    precioServicio = float('%.2f' % precioServicio)
                if serviceItem[3] == 'DOLARES':
                    precioServicio = Decimal(serviceItem[4])*Decimal(invoiceObject.erSel)*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                    precioServicio = float('%.2f' % precioServicio)
                infoServiceItem = {
                    "codigoProducto":serviceItem[0],
                    "codigoProductoSunat":"",
                    "descripcion":serviceItem[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_BIENES",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precioServicio,
                    "descuento":{
                        "monto":'0',
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                finalPrice = finalPrice + precioServicio
                itemsInvoice.append(infoServiceItem)
                i = i + 1
        else:
            i = 1
            for serviceItem in totalServicesItems:
                if serviceItem[3] == 'DOLARES':
                    precioServicio = Decimal(serviceItem[4])*Decimal(Decimal(1.00) - Decimal(serviceItem[5])/100)
                    precioServicio = float('%.2f' % precioServicio)
                if serviceItem[3] == 'SOLES':
                    precioServicio = (Decimal(serviceItem[4])/Decimal(invoiceObject.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                    precioServicio = float('%.2f' % precioServicio)
                infoServiceItem = {
                    "codigoProducto":serviceItem[0],
                    "codigoProductoSunat":"",
                    "descripcion":serviceItem[1],
                    "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                    "unidadMedida":"UNIDAD_BIENES",
                    "cantidad":1,
                    "valorVentaUnitarioItem":precioServicio,
                    "descuento":{
                        "monto":'0',
                    },
                    "numeroOrden":i,
                    "esPorcentaje":True
                }
                finalPrice = finalPrice + precioServicio
                itemsInvoice.append(infoServiceItem)
                i = i + 1
    
    quotesData = []
    if condicionPago == 'CONTADO':
        quotesData = None
    else:
        totalQuotes = len(invoiceObject.dateQuotesInvoice)
        finalPrice = float('%.2f' % finalPrice)
        monto = '%.2f' % ((finalPrice*1.18)/int(totalQuotes))
        i=0
        while i < totalQuotes:
            quoteInfo = {
                "numero":'00' + str(i+1),
                "fecha":invoiceObject.dateQuotesInvoice[i],
                "monto":str(monto),
                "moneda":currencyInvoice
            }
            i = i+1
            quotesData.append(quoteInfo)
    
    if len(documentClient) == 11:
        tipoDocumentoBoleta = 'RUC'
    else:
        tipoDocumentoBoleta = 'DOC_NACIONAL_DE_IDENTIDAD'
    
    param_data = {
        "close2u":refSuperior, 
        "datosDocumento":
        {
            "serie":invoiceObject.endpointInvoice.serieBoleta,
            "numero":invoiceObject.nroInvoice,
            "moneda":currencyInvoice,
            "fechaEmision":invoiceObject.dateInvoice.strftime("%Y-%m-%d"),
            "horaEmision":None,
            "fechaVencimiento": None,
            "formaPago":condicionPago,
            "ordencompra":invoiceObject.relatedDocumentInvoice,
            "puntoEmisor":None,
            "glosa":str(invoiceObject.commentInvoice),
        },
        "detalleDocumento":itemsInvoice,
        "emisor":
        {
            "correo":"info@metalprotec.com",
            "nombreComercial": None,
            "nombreLegal": "METALPROTEC",
            "numeroDocumentoIdentidad": "20541628631",
            "tipoDocumentoIdentidad": "RUC",
            "domicilioFiscal":
            {
                "pais":"PERU",
                "departamento":"ANCASH",
                "provincia":"SANTA",
                "distrito":"NUEVO CHIMBOTE",
                "direccon":"Mza. J4 Lote. 39",
                "ubigeo":"02712"
            }
        },
        "informacionAdicional":
        {
            "tipoOperacion":"VENTA_INTERNA",
            "coVendedor":None,
            "vendedor":sellerInfo
        },
        "receptor":
        {
            "correo":emailClient,
            "correoCopia":None,
            "domicilioFiscal":
            {
                "departamento":None,
                "direccion":addressClient,
                "distrito":None,
                "pais":"PERU",
                "provincia": None,
                "ubigeo": None,
                "urbanizacion": None
            },
            "nombreComercial":None,
            "nombreLegal":identificationClient,
            "numeroDocumentoIdentidad":documentClient,
            "tipoDocumentoIdentidad":tipoDocumentoBoleta
        },
        'referencias':{
            "documentoReferenciaList":guideJson,
        },
        'cuotas':quotesData
    }
    return param_data

def getInfoCreditNote(creditNoteObject):
    if creditNoteObject.originCreditNote == 'INVOICE':
        invoiceObject = creditNoteObject.asociatedInvoice
        if invoiceObject.typeItemsInvoice == 'SERVICIOS':
            refSuperior = {
                "tipoIntegracion":"OFFLINE",
                "tipoPlantilla":"01",
                "tipoRegistro":"PRECIOS_SIN_IGV"
            }
        else:
            refSuperior = {
                "tipoIntegracion":"OFFLINE",
                "tipoPlantilla":"01",
                "tipoRegistro":"PRECIOS_SIN_IGV"
            }

        if invoiceObject.currencyInvoice == 'DOLARES':
            currencyInvoice = 'USD'
        else:
            currencyInvoice = 'PEN'

        if invoiceObject.originInvoice == 'GUIDE':
            emailClient = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[4]
            addressClient = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[7]
            identificationClient = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[1]
            documentClient = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[2]
            condicionPago = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.paymentQuotation
            sellerInfo = invoiceObject.guidesystem_set.all()[0].asociatedQuotation.quotationsellerdata.dataUserQuotation[1]
            guideJson = []
            for guideInfo in invoiceObject.guidesystem_set.all():
                newGuide = {
                    "tipoDocumento":"GUIAEMISIONREMITENTE",
                    "numero":guideInfo.nroGuide,
                    "serie":invoiceObject.endpointInvoice.serieGuia,
                }
                guideJson.append(newGuide)
        else:
            emailClient = invoiceObject.asociatedQuotation.quotationclientdata.dataClientQuotation[4]
            addressClient = invoiceObject.asociatedQuotation.quotationclientdata.dataClientQuotation[7]
            identificationClient = invoiceObject.asociatedQuotation.quotationclientdata.dataClientQuotation[1]
            documentClient = invoiceObject.asociatedQuotation.quotationclientdata.dataClientQuotation[2]
            condicionPago = invoiceObject.asociatedQuotation.paymentQuotation
            sellerInfo = invoiceObject.asociatedQuotation.quotationsellerdata.dataUserQuotation[1]
            guideJson = None

        itemsInvoice = []
        if invoiceObject.typeItemsInvoice == 'PRODUCTOS':
            totalProductsItems = []
            if invoiceObject.originInvoice == 'GUIDE':
                for guideInfo in invoiceObject.guidesystem_set.all():
                    for productInfo in guideInfo.asociatedQuotation.quotationproductdata_set.all():
                        if productInfo.asociatedProduct is not None:
                            if productInfo.asociatedProduct.kitProduct == 'ON':
                                productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                                totalProductsItems.append([
                                    productoDatos.id,
                                    productoDatos.nameProduct,
                                    productoDatos.codeProduct,
                                    productoDatos.measureUnit,
                                    productInfo.dataProductQuotation[4],
                                    productInfo.dataProductQuotation[5],
                                    productInfo.dataProductQuotation[6],
                                    productInfo.dataProductQuotation[7],
                                    productInfo.dataProductQuotation[8],
                                    productInfo.dataProductQuotation[9],
                                ])
                            else:
                                totalProductsItems.append(productInfo.dataProductQuotation)
                        else:
                            totalProductsItems.append(productInfo.dataProductQuotation)
            else:
                for productInfo in invoiceObject.asociatedQuotation.quotationproductdata_set.all():
                    if productInfo.asociatedProduct is not None:
                        if productInfo.asociatedProduct.kitProduct == 'ON':
                            productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                            totalProductsItems.append([
                                productoDatos.id,
                                productoDatos.nameProduct,
                                productoDatos.codeProduct,
                                productoDatos.measureUnit,
                                productInfo.dataProductQuotation[4],
                                productInfo.dataProductQuotation[5],
                                productInfo.dataProductQuotation[6],
                                productInfo.dataProductQuotation[7],
                                productInfo.dataProductQuotation[8],
                                productInfo.dataProductQuotation[9],
                            ])
                        else:
                            totalProductsItems.append(productInfo.dataProductQuotation)
                    else:
                        totalProductsItems.append(productInfo.dataProductQuotation)
            
            finalPrice = 0
            if currencyInvoice == 'PEN':
                i = 1
                for productoItem in totalProductsItems:
                    if productoItem[5] == 'SOLES':
                        precioProducto = Decimal(productoItem[6])*Decimal(Decimal(1.00) - (Decimal(productoItem[7])/100))
                        precioProducto = float('%.2f' % precioProducto)
                    else:
                        precioProducto = Decimal(productoItem[6])*Decimal(invoiceObject.erSel)*Decimal(Decimal(1.00) - Decimal(productoItem[7])/100)
                        precioProducto = float('%.2f' % precioProducto)
                    if productoItem[9] == '1':
                        infoProductoItem = {
                            "codigoProducto":productoItem[2],
                            "codigoProductoSunat":"",
                            "descripcion":productoItem[1],
                            "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                            "unidadMedida":"UNIDAD_BIENES",
                            "cantidad":str(int(float(precioProducto[8]))),
                            "valorReferencialUnitarioItem":precioProducto,
                            "descuento":{
                                "monto":'0',
                            },
                            "numeroOrden":i,
                            "esPorcentaje":True
                        }
                        finalPrice = finalPrice
                    else:
                        infoProductoItem = {
                            "codigoProducto":productoItem[2],
                            "codigoProductoSunat":"",
                            "descripcion":productoItem[1],
                            "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                            "unidadMedida":"UNIDAD_BIENES",
                            "cantidad":str(int(float(productoItem[8]))),
                            "valorVentaUnitarioItem":precioProducto,
                            "descuento":{
                                "monto":'0',
                            },
                            "numeroOrden":i,
                            "esPorcentaje":True
                        }
                        finalPrice = finalPrice + precioProducto*round(float(productoItem[8]),2)
                    itemsInvoice.append(infoProductoItem)
                    i = i +1
            else:
                i = 1
                for productoItem in totalProductsItems:
                    if productoItem[5] == 'SOLES':
                        precioProducto = (Decimal(productoItem[6])/Decimal(invoiceObject.erSel))*Decimal(Decimal(1.00) - (Decimal(productoItem[7])/100))
                        precioProducto = float('%.2f' % precioProducto)
                    else:
                        precioProducto = Decimal(productoItem[6])*Decimal(Decimal(1.00) - Decimal(productoItem[7])/100)
                        precioProducto = float('%.2f' % precioProducto)
                    if productoItem[9] == '1':
                        infoProductoItem = {
                            "codigoProducto":productoItem[2],
                            "codigoProductoSunat":"",
                            "descripcion":productoItem[1],
                            "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                            "unidadMedida":"UNIDAD_BIENES",
                            "cantidad":str(int(float(precioProducto[8]))),
                            "valorReferencialUnitarioItem":precioProducto,
                            "descuento":{
                                "monto":'0',
                            },
                            "numeroOrden":i,
                            "esPorcentaje":True
                        }
                        finalPrice = finalPrice
                    else:
                        infoProductoItem = {
                            "codigoProducto":productoItem[2],
                            "codigoProductoSunat":"",
                            "descripcion":productoItem[1],
                            "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                            "unidadMedida":"UNIDAD_BIENES",
                            "cantidad":str(int(float(productoItem[8]))),
                            "valorVentaUnitarioItem":precioProducto,
                            "descuento":{
                                "monto":'0',
                            },
                            "numeroOrden":i,
                            "esPorcentaje":True
                        }
                        finalPrice = finalPrice + precioProducto*round(float(productoItem[8]),2)
                    itemsInvoice.append(infoProductoItem)
                    i = i +1
        else:
            totalServicesItems = []
            for serviceInfo in invoiceObject.asociatedQuotation.quotationservicedata_set.all():
                totalServicesItems.append(serviceInfo.dataServiceQuotation)
            finalPrice = 0
            if currencyInvoice == 'PEN':
                i = 1
                for serviceItem in totalServicesItems:
                    if serviceItem[3] == 'SOLES':
                        precioServicio = Decimal(serviceItem[4])*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                        precioServicio = float('%.2f' % precioServicio)
                    if serviceItem[3] == 'DOLARES':
                        precioServicio = Decimal(serviceItem[4])*Decimal(invoiceObject.erSel)*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                        precioServicio = float('%.2f' % precioServicio)
                    infoServiceItem = {
                        "codigoProducto":serviceItem[0],
                        "codigoProductoSunat":"",
                        "descripcion":serviceItem[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":1,
                        "valorVentaUnitarioItem":precioServicio,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice + precioServicio
                    itemsInvoice.append(infoServiceItem)
                    i = i + 1
            else:
                i = 1
                for serviceItem in totalServicesItems:
                    if serviceItem[3] == 'DOLARES':
                        precioServicio = Decimal(serviceItem[4])*Decimal(Decimal(1.00) - Decimal(serviceItem[5])/100)
                        precioServicio = float('%.2f' % precioServicio)
                    if serviceItem[3] == 'SOLES':
                        precioServicio = (Decimal(serviceItem[4])/Decimal(invoiceObject.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                        precioServicio = float('%.2f' % precioServicio)
                    infoServiceItem = {
                        "codigoProducto":serviceItem[0],
                        "codigoProductoSunat":"",
                        "descripcion":serviceItem[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":1,
                        "valorVentaUnitarioItem":precioServicio,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice + precioServicio
                    itemsInvoice.append(infoServiceItem)
                    i = i + 1
        
        quotesData = []
        if condicionPago == 'CONTADO':
            quotesData = None
        else:
            totalQuotes = len(invoiceObject.dateQuotesInvoice)
            finalPrice = float('%.2f' % finalPrice)
            monto = '%.2f' % ((finalPrice*1.18)/int(totalQuotes))
            i=0
            while i < totalQuotes:
                quoteInfo = {
                    "numero":'00' + str(i+1),
                    "fecha":invoiceObject.dateQuotesInvoice[i],
                    "monto":str(monto),
                    "moneda":currencyInvoice
                }
                i = i+1
                quotesData.append(quoteInfo)
        
        if len(documentClient) == 11:
            tipoDocumentoBoleta = 'RUC'
        else:
            tipoDocumentoBoleta = 'DOC_NACIONAL_DE_IDENTIDAD'

        nroComprobante = int(invoiceObject.nroInvoice)
        serieComprobante = invoiceObject.endpointInvoice.serieBoleta
        typeDocument = 'BOLETA'
        dateDocument = invoiceObject.dateInvoice.strftime("%Y-%m-%d")
    else:
        billObject = creditNoteObject.asociatedBill
        if billObject.typeItemsBill == 'SERVICIOS':
            refSuperior = {
                "tipoIntegracion":"OFFLINE",
                "tipoPlantilla":"01",
                "tipoRegistro":"PRECIOS_SIN_IGV"
            }
        else:
            refSuperior = {
                "tipoIntegracion":"OFFLINE",
                "tipoPlantilla":"01",
                "tipoRegistro":"PRECIOS_SIN_IGV"
            }

        if billObject.currencyBill == 'DOLARES':
            currencyBill = 'USD'
        else:
            currencyBill = 'PEN'

        if billObject.originBill == 'GUIDE':
            emailClient = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[4]
            addressClient = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[7]
            identificationClient = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[1]
            documentClient = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationclientdata.dataClientQuotation[2]
            condicionPago = billObject.guidesystem_set.all()[0].asociatedQuotation.paymentQuotation
            sellerInfo = billObject.guidesystem_set.all()[0].asociatedQuotation.quotationsellerdata.dataUserQuotation[1]
            guideJson = []
            for guideInfo in billObject.guidesystem_set.all():
                newGuide = {
                    "tipoDocumento":"GUIAEMISIONREMITENTE",
                    "numero":guideInfo.nroGuide,
                    "serie":billObject.endpointBill.serieGuia,
                }
                guideJson.append(newGuide)
        else:
            emailClient = billObject.asociatedQuotation.quotationclientdata.dataClientQuotation[4]
            addressClient = billObject.asociatedQuotation.quotationclientdata.dataClientQuotation[7]
            identificationClient = billObject.asociatedQuotation.quotationclientdata.dataClientQuotation[1]
            documentClient = billObject.asociatedQuotation.quotationclientdata.dataClientQuotation[2]
            condicionPago = billObject.asociatedQuotation.paymentQuotation
            sellerInfo = billObject.asociatedQuotation.quotationsellerdata.dataUserQuotation[1]
            guideJson = None

        itemsBill = []
        if billObject.typeItemsBill == 'PRODUCTOS':
            totalProductsItems = []
            if creditNoteObject.creditNotePurpose == 'ANULACION_OPERACION':
                if billObject.originBill == 'GUIDE':
                    for guideInfo in billObject.guidesystem_set.all():
                        for productInfo in guideInfo.asociatedQuotation.quotationproductdata_set.all():
                            if productInfo.asociatedProduct is not None:
                                if productInfo.asociatedProduct.kitProduct == 'ON':
                                    productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                                    totalProductsItems.append([
                                        productoDatos.id,
                                        productoDatos.nameProduct,
                                        productoDatos.codeProduct,
                                        productoDatos.measureUnit,
                                        productInfo.dataProductQuotation[4],
                                        productInfo.dataProductQuotation[5],
                                        productInfo.dataProductQuotation[6],
                                        productInfo.dataProductQuotation[7],
                                        productInfo.dataProductQuotation[8],
                                        productInfo.dataProductQuotation[9],
                                    ])
                                else:
                                    totalProductsItems.append(productInfo.dataProductQuotation)
                            else:
                                totalProductsItems.append(productInfo.dataProductQuotation)
                else:
                    for productInfo in billObject.asociatedQuotation.quotationproductdata_set.all():
                        if productInfo.asociatedProduct is not None:
                            if productInfo.asociatedProduct.kitProduct == 'ON':
                                productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                                totalProductsItems.append([
                                    productoDatos.id,
                                    productoDatos.nameProduct,
                                    productoDatos.codeProduct,
                                    productoDatos.measureUnit,
                                    productInfo.dataProductQuotation[4],
                                    productInfo.dataProductQuotation[5],
                                    productInfo.dataProductQuotation[6],
                                    productInfo.dataProductQuotation[7],
                                    productInfo.dataProductQuotation[8],
                                    productInfo.dataProductQuotation[9],
                                ])
                            else:
                                totalProductsItems.append(productInfo.dataProductQuotation)
                        else:
                            totalProductsItems.append(productInfo.dataProductQuotation)
            else:
                if billObject.originBill == 'GUIDE':
                    for guideInfo in billObject.guidesystem_set.all():
                        for productInfo in guideInfo.asociatedQuotation.quotationproductdata_set.all():
                            if productInfo.dataProductQuotation[2] in creditNoteObject.codigosProductos:
                                indexCodeQuantity = creditNoteObject.codigosProductos.index(productInfo.dataProductQuotation[2])
                                productInfo.dataProductQuotation[8] = creditNoteObject.cantidadesProductos[indexCodeQuantity]
                                if productInfo.asociatedProduct is not None:
                                    if productInfo.asociatedProduct.kitProduct == 'ON':
                                        productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                                        totalProductsItems.append([
                                            productoDatos.id,
                                            productoDatos.nameProduct,
                                            productoDatos.codeProduct,
                                            productoDatos.measureUnit,
                                            productInfo.dataProductQuotation[4],
                                            productInfo.dataProductQuotation[5],
                                            productInfo.dataProductQuotation[6],
                                            productInfo.dataProductQuotation[7],
                                            productInfo.dataProductQuotation[8],
                                            productInfo.dataProductQuotation[9],
                                        ])
                                    else:
                                        totalProductsItems.append(productInfo.dataProductQuotation)
                                else:
                                    totalProductsItems.append(productInfo.dataProductQuotation)
                else:
                    for productInfo in billObject.asociatedQuotation.quotationproductdata_set.all():
                        if productInfo.dataProductQuotation[2] in creditNoteObject.codigosProductos:
                            indexCodeQuantity = creditNoteObject.codigosProductos.index(productInfo.dataProductQuotation[2])
                            productInfo.dataProductQuotation[8] = creditNoteObject.cantidadesProductos[indexCodeQuantity]
                            if productInfo.asociatedProduct is not None:
                                if productInfo.asociatedProduct.kitProduct == 'ON':
                                    productoDatos = productSystem.objects.get(id=productInfo.asociatedProduct.kitInfo[0][0])
                                    totalProductsItems.append([
                                        productoDatos.id,
                                        productoDatos.nameProduct,
                                        productoDatos.codeProduct,
                                        productoDatos.measureUnit,
                                        productInfo.dataProductQuotation[4],
                                        productInfo.dataProductQuotation[5],
                                        productInfo.dataProductQuotation[6],
                                        productInfo.dataProductQuotation[7],
                                        productInfo.dataProductQuotation[8],
                                        productInfo.dataProductQuotation[9],
                                    ])
                                else:
                                    totalProductsItems.append(productInfo.dataProductQuotation)
                            else:
                                totalProductsItems.append(productInfo.dataProductQuotation)
            
            finalPrice = 0
            if currencyBill == 'PEN':
                i = 1
                for productoItem in totalProductsItems:
                    if productoItem[5] == 'SOLES':
                        precioProducto = Decimal(productoItem[6])*Decimal(Decimal(1.00) - (Decimal(productoItem[7])/100))
                        precioProducto = float('%.2f' % precioProducto)
                    else:
                        precioProducto = Decimal(productoItem[6])*Decimal(billObject.erSel)*Decimal(Decimal(1.00) - Decimal(productoItem[7])/100)
                        precioProducto = float('%.2f' % precioProducto)
                    if productoItem[9] == '1':
                        infoProductoItem = {
                            "codigoProducto":productoItem[2],
                            "codigoProductoSunat":"",
                            "descripcion":productoItem[1],
                            "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                            "unidadMedida":"UNIDAD_BIENES",
                            "cantidad":str(int(float(precioProducto[8]))),
                            "valorReferencialUnitarioItem":precioProducto,
                            "descuento":{
                                "monto":'0',
                            },
                            "numeroOrden":i,
                            "esPorcentaje":True
                        }
                        finalPrice = finalPrice
                    else:
                        infoProductoItem = {
                            "codigoProducto":productoItem[2],
                            "codigoProductoSunat":"",
                            "descripcion":productoItem[1],
                            "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                            "unidadMedida":"UNIDAD_BIENES",
                            "cantidad":str(int(float(productoItem[8]))),
                            "valorVentaUnitarioItem":precioProducto,
                            "descuento":{
                                "monto":'0',
                            },
                            "numeroOrden":i,
                            "esPorcentaje":True
                        }
                        finalPrice = finalPrice + precioProducto*round(float(productoItem[8]),2)
                    itemsBill.append(infoProductoItem)
                    i = i +1
            else:
                i = 1
                for productoItem in totalProductsItems:
                    if productoItem[5] == 'SOLES':
                        precioProducto = (Decimal(productoItem[6])/Decimal(billObject.erSel))*Decimal(Decimal(1.00) - (Decimal(productoItem[7])/100))
                        precioProducto = float('%.2f' % precioProducto)
                    else:
                        precioProducto = Decimal(productoItem[6])*Decimal(Decimal(1.00) - Decimal(productoItem[7])/100)
                        precioProducto = float('%.2f' % precioProducto)
                    if productoItem[9] == '1':
                        infoProductoItem = {
                            "codigoProducto":productoItem[2],
                            "codigoProductoSunat":"",
                            "descripcion":productoItem[1],
                            "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                            "unidadMedida":"UNIDAD_BIENES",
                            "cantidad":str(int(float(precioProducto[8]))),
                            "valorReferencialUnitarioItem":precioProducto,
                            "descuento":{
                                "monto":'0',
                            },
                            "numeroOrden":i,
                            "esPorcentaje":True
                        }
                        finalPrice = finalPrice
                    else:
                        infoProductoItem = {
                            "codigoProducto":productoItem[2],
                            "codigoProductoSunat":"",
                            "descripcion":productoItem[1],
                            "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                            "unidadMedida":"UNIDAD_BIENES",
                            "cantidad":str(int(float(productoItem[8]))),
                            "valorVentaUnitarioItem":precioProducto,
                            "descuento":{
                                "monto":'0',
                            },
                            "numeroOrden":i,
                            "esPorcentaje":True
                        }
                        finalPrice = finalPrice + precioProducto*round(float(productoItem[8]),2)
                    itemsBill.append(infoProductoItem)
                    i = i +1
        else:
            totalServicesItems = []
            for serviceInfo in billObject.asociatedQuotation.quotationservicedata_set.all():
                totalServicesItems.append(serviceInfo.dataServiceQuotation)
            finalPrice = 0
            if currencyBill == 'PEN':
                i = 1
                for serviceItem in totalServicesItems:
                    if serviceItem[3] == 'SOLES':
                        precioServicio = Decimal(serviceItem[4])*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                        precioServicio = float('%.2f' % precioServicio)
                    if serviceItem[3] == 'DOLARES':
                        precioServicio = Decimal(serviceItem[4])*Decimal(billObject.erSel)*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                        precioServicio = float('%.2f' % precioServicio)
                    infoServiceItem = {
                        "codigoProducto":serviceItem[0],
                        "codigoProductoSunat":"",
                        "descripcion":serviceItem[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":1,
                        "valorVentaUnitarioItem":precioServicio,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice + precioServicio
                    itemsBill.append(infoServiceItem)
                    i = i + 1
            else:
                i = 1
                for serviceItem in totalServicesItems:
                    if serviceItem[3] == 'DOLARES':
                        precioServicio = Decimal(serviceItem[4])*Decimal(Decimal(1.00) - Decimal(serviceItem[5])/100)
                        precioServicio = float('%.2f' % precioServicio)
                    if serviceItem[3] == 'SOLES':
                        precioServicio = (Decimal(serviceItem[4])/Decimal(billObject.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceItem[5])/100))
                        precioServicio = float('%.2f' % precioServicio)
                    infoServiceItem = {
                        "codigoProducto":serviceItem[0],
                        "codigoProductoSunat":"",
                        "descripcion":serviceItem[1],
                        "tipoAfectacion":"GRAVADO_OPERACION_ONEROSA",
                        "unidadMedida":"UNIDAD_BIENES",
                        "cantidad":1,
                        "valorVentaUnitarioItem":precioServicio,
                        "descuento":{
                            "monto":'0',
                        },
                        "numeroOrden":i,
                        "esPorcentaje":True
                    }
                    finalPrice = finalPrice + precioServicio
                    itemsBill.append(infoServiceItem)
                    i = i + 1
        
        quotesData = []
        if condicionPago == 'CONTADO':
            quotesData = None
        else:
            totalQuotes = len(billObject.dateQuotesBill)
            finalPrice = float('%.2f' % finalPrice)
            monto = '%.2f' % ((finalPrice*1.18)/int(totalQuotes))
            i=0
            while i < totalQuotes:
                quoteInfo = {
                    "numero":'00' + str(i+1),
                    "fecha":billObject.dateQuotesBill[i],
                    "monto":str(monto),
                    "moneda":currencyBill
                }
                i = i+1
                quotesData.append(quoteInfo)
        nroComprobante = int(billObject.nroBill)
        serieComprobante = billObject.endpointBill.serieFactura
        typeDocument = 'FACTURA'
        dateDocument = billObject.dateBill.strftime("%Y-%m-%d")

    if creditNoteObject.originCreditNote == 'INVOICE':
        currencyNote = currencyInvoice
        serieNota = invoiceObject.endpointInvoice.serieNotaBoleta
        documentoIdentidad = tipoDocumentoBoleta
        itemsNote = itemsInvoice
    else:
        currencyNote = currencyBill
        serieNota = billObject.endpointBill.serieNotaFactura
        documentoIdentidad = 'RUC'
        itemsNote = itemsBill

    param_data = {
        "close2u":refSuperior,
        "comprobanteAjustado":{
            "serie":serieComprobante,
            "numero":nroComprobante,
            "tipoDocumento":typeDocument,
            "fechaEmision":dateDocument
        },
        "datosDocumento":
        {
            "serie":serieNota,
            "numero":creditNoteObject.nroCreditNote,
            "moneda":currencyNote,
            "fechaEmision":creditNoteObject.dateCreditNote.strftime("%Y-%m-%d"),
            "horaEmision":None,
            "fechaVencimiento": None,
            "formaPago":"CONTADO",
            "medioPago": "DEPOSITO_CUENTA",
            "condicionPago": None,
            "ordencompra":None,
            "puntoEmisor":None,
            "glosa":creditNoteObject.creditNotePurpose,
        },
        "detalleDocumento":itemsNote,
        "emisor":
        {
            "correo":"info@metalprotec.com",
            "nombreComercial": None,
            "nombreLegal": "METALPROTEC",
            "numeroDocumentoIdentidad": "20541628631",
            "tipoDocumentoIdentidad": "RUC",
            "domicilioFiscal":
            {
                "pais":"PERU",
                "departamento":"ANCASH",
                "provincia":"SANTA",
                "distrito":"NUEVO CHIMBOTE",
                "direccon":"Mza. J4 Lote. 39",
                "ubigeo":"02712"
            }
        },
        "informacionAdicional":
        {
            "tipoOperacion":"VENTA_INTERNA",
            "coVendedor":None,
        },
        "motivo":creditNoteObject.creditNotePurpose,
        "receptor":
        {
            "correo":emailClient,
            "correoCopia":None,
            "domicilioFiscal":
            {
                "direccion":addressClient,
                "pais":"PERU",
            },
            "nombreComercial":None,
            "nombreLegal":identificationClient,
            "numeroDocumentoIdentidad":documentClient,
            "tipoDocumentoIdentidad":documentoIdentidad
        },
        'cuotas':None
    }
    return param_data


def createCreditNoteFromBill(request):
    if request.method == 'POST':
        idBill = request.POST.get('idBill')
        creditNotePurpose = request.POST.get('creditNotePurpose')
        if creditNotePurpose == 'ANULACION_OPERACION':
            asociatedBill = billSystem.objects.get(id=idBill)
            endpointCreditNote = request.user.extendeduser.endpointUser
            serieCreditNote = endpointCreditNote.serieNotaFactura
            nroCreditNote = endpointCreditNote.nroNotaFactura
            endpointCreditNote.nroNotaFactura = str(int(nroCreditNote) + 1)
            endpointCreditNote.save()
            codeCreditNote = nroCreditNote
            while len(codeCreditNote) < 4:
                codeCreditNote = '0' + codeCreditNote
            codeCreditNote = f"{serieCreditNote}-{codeCreditNote}"
            creditNoteSystem.objects.create(
                asociatedBill=asociatedBill,
                typeCreditNote='BILL',
                stateCreditNote='GENERADA',
                codeCreditNote=codeCreditNote,
                stateTeFacturo='',
                dateCreditNote=datetime.datetime.today(),
                nroCreditNote=nroCreditNote,
                originCreditNote='BILL',
                creditNotePurpose=creditNotePurpose,
                endpointCreditNote=endpointCreditNote
            )
            return HttpResponseRedirect(reverse('salesMetalprotec:creditNotesMetalprotec'))
        else:
            codigosProductos = request.POST.getlist('codigoProducto')
            cantidadesProductos = request.POST.getlist('cantidadProducto')
            asociatedBill = billSystem.objects.get(id=idBill)
            endpointCreditNote = request.user.extendeduser.endpointUser
            serieCreditNote = endpointCreditNote.serieNotaFactura
            nroCreditNote = endpointCreditNote.nroNotaFactura
            endpointCreditNote.nroNotaFactura = str(int(nroCreditNote) + 1)
            endpointCreditNote.save()
            codeCreditNote = nroCreditNote
            while len(codeCreditNote) < 4:
                codeCreditNote = '0' + codeCreditNote
            codeCreditNote = f"{serieCreditNote}-{codeCreditNote}"
            creditNoteSystem.objects.create(
                asociatedBill=asociatedBill,
                typeCreditNote='BILL',
                stateCreditNote='GENERADA',
                codeCreditNote=codeCreditNote,
                stateTeFacturo='',
                dateCreditNote=datetime.datetime.today(),
                nroCreditNote=nroCreditNote,
                originCreditNote='BILL',
                creditNotePurpose=creditNotePurpose,
                codigosProductos=codigosProductos,
                cantidadesProductos=cantidadesProductos,
                endpointCreditNote=endpointCreditNote
            )
            return HttpResponseRedirect(reverse('salesMetalprotec:creditNotesMetalprotec'))


def createCreditNoteFromInvoice(request,idInvoice):
    asociatedInvoice = invoiceSystem.objects.create(id=idInvoice)
    endpointCreditNote = request.user.extendeduser.endpointUser
    serieCreditNote = endpointCreditNote.serieNotaBoleta
    nroCreditNote = endpointCreditNote.nroNotaBoleta
    endpointCreditNote.nroNotaBoleta = str(int(nroCreditNote) + 1)
    endpointCreditNote.save()
    codeCreditNote = nroCreditNote
    while len(codeCreditNote) < 4:
        codeCreditNote = '0' + codeCreditNote
    codeCreditNote = f"{serieCreditNote}-{codeCreditNote}"
    creditNoteSystem.objects.create(
        asociatedInvoice=asociatedInvoice,
        typeCreditNote='INVOICE',
        stateCreditNote='GENERADA',
        codeCreditNote=codeCreditNote,
        stateTeFacturo='',
        dateCreditNote=datetime.datetime.today(),
        nroCreditNote=nroCreditNote,
        originCreditNote='INVOICE',
        endpointCreditNote=endpointCreditNote
    )
    return HttpResponseRedirect(reverse('salesMetalprotec:creditNotesMetalprotec'))


def createBillFromGuides(request):
    if request.method == 'POST':
        guidesData = json.load(request)
        guidesInfo = guidesData.get('guidesInfo')
        probeGuide = guideSystem.objects.get(id=guidesInfo[0])

        endpointBill = request.user.extendeduser.endpointUser
        nroBill = endpointBill.nroFactura
        serieBill = endpointBill.serieFactura
        endpointBill.nroFactura = str(int(nroBill) + 1)
        endpointBill.save()

        codeBill = nroBill
        while len(codeBill) < 4:
            codeBill = '0' + codeBill
        codeBill = f"{serieBill}-{codeBill}"

        dateQuotesBill = []
        if probeGuide.asociatedQuotation.paymentQuotation == 'CONTADO':
            dateQuotesBill = []
        else:
            for numberData in range(int(probeGuide.asociatedQuotation.quotesQuotation)):
                dateQuotesBill.append('2023-01-01')
        infoCreatedBill = billSystem.objects.create(
            commentBill='',
            dateBill=datetime.datetime.today(),
            relatedDocumentBill='',
            dateQuotesBill=dateQuotesBill,
            erBuy=probeGuide.asociatedQuotation.erBuy,
            erSel=probeGuide.asociatedQuotation.erSel,
            currencyBill=probeGuide.asociatedQuotation.currencyQuotation,
            stateBill='GENERADA',
            codeBill=codeBill,
            stateTeFacturo='',
            nroBill=nroBill,
            typeItemsBill='PRODUCTOS',
            originBill='GUIDE',
            endpointBill=endpointBill
        )

        for idGuide in guidesInfo:
            guideObject = guideSystem.objects.get(id=idGuide)
            guideObject.asociatedBill = infoCreatedBill
            guideObject.save()
        return HttpResponseRedirect(reverse('salesMetalprotec:billsMetalprotec'))
    
def createInvoiceFromGuides(request):
    if request.method == 'POST':
        guidesData = json.load(request)
        guidesInfo = guidesData.get('guidesInfo')
        probeGuide = guideSystem.objects.get(id=guidesInfo[0])

        endpointInvoice = request.user.extendeduser.endpointUser
        nroInvoice = endpointInvoice.nroBoleta
        serieInvoice = endpointInvoice.serieBoleta
        endpointInvoice.nroBoleta = str(int(nroInvoice) + 1)
        endpointInvoice.save()

        codeInvoice = nroInvoice
        while len(codeInvoice) < 4:
            codeInvoice = '0' + codeInvoice
        codeInvoice = f"{serieInvoice}-{codeInvoice}"

        dateQuotesInvoice = []
        if probeGuide.asociatedQuotation.paymentQuotation == 'CONTADO':
            dateQuotesInvoice = []
        else:
            for numberData in range(int(probeGuide.asociatedQuotation.quotesQuotation)):
                dateQuotesInvoice.append('2023-01-01')
        infoCreatedInvoice = invoiceSystem.objects.create(
            commentInvoice='',
            dateInvoice=datetime.datetime.today(),
            relatedDocumentInvoice='',
            dateQuotesInvoice=dateQuotesInvoice,
            erBuy=probeGuide.asociatedQuotation.erBuy,
            erSel=probeGuide.asociatedQuotation.erSel,
            currencyInvoice=probeGuide.asociatedQuotation.currencyQuotation,
            stateInvoice='GENERADA',
            codeInvoice=codeInvoice,
            stateTeFacturo='',
            nroInvoice=nroInvoice,
            typeItemsInvoice='PRODUCTOS',
            originInvoice='GUIDE',
            endpointInvoice=endpointInvoice
        )

        for idGuide in guidesInfo:
            guideObject = guideSystem.objects.get(id=idGuide)
            guideObject.asociatedInvoice = infoCreatedInvoice
            guideObject.save()
        return HttpResponseRedirect(reverse('salesMetalprotec:invoicesMetalprotec'))
    
def getExchangeRate():
    exchangeRate = []
    try:
        r = requests.get('https://www.sbs.gob.pe/app/pp/sistip_portal/paginas/publicacion/tipocambiopromedio.aspx')
        datos = BeautifulSoup(r.text,'html.parser')
        tc_fila = datos.find(id='ctl00_cphContent_rgTipoCambio_ctl00__0')
        tc_fila = tc_fila.find_all(class_='APLI_fila2')
        if len(tc_fila) == 2:
            tc_compra = round(float(tc_fila[0].string),3)
            tc_venta = round(float(tc_fila[1].string),3)
            exchangeRate = []
            exchangeRate.append(str(tc_compra))
            exchangeRate.append(str(tc_venta))
        else:
            tc_compra = 0.000
            tc_venta = 0.000
            exchangeRate = []
            exchangeRate.append(str(tc_compra))
            exchangeRate.append(str(tc_venta))
    except:
        tc_compra = 3.705
        tc_venta = 3.710
        exchangeRate = []
        exchangeRate.append(str(tc_compra))
        exchangeRate.append(str(tc_venta))
    return exchangeRate

def exportFilteredQuotations(request):
    if request.method == 'POST':
        startDate = request.POST.get('startDate')
        endDate = request.POST.get('endDate')
        quotationData = []
        if startDate != '' and endDate != '':
            fechaInicio = datetime.datetime.strptime(startDate,'%Y-%m-%d').date()
            fechaFin = datetime.datetime.strptime(endDate,'%Y-%m-%d').date()
            quotationFilter = quotationSystem.objects.filter(
                Q(dateQuotation__gte=fechaInicio) &
                Q(dateQuotation__lte=fechaFin)
            ).order_by('-dateQuotation')
            for quotationItem in quotationFilter:
                quotationData.append([
                    quotationItem.dateQuotation.strftime('%Y-%m-%d'),
                    quotationItem.codeQuotation,
                    quotationItem.quotationclientdata.dataClientQuotation[1],
                    quotationItem.stateQuotation,
                    quotationItem.currencyQuotation,
                    getValueQuotation(quotationItem),
                    getSolesValue(quotationItem),
                ])
            finalPrice = Decimal(0.00)
            for itemInfo in quotationData:
                finalPrice = Decimal(finalPrice) + Decimal(itemInfo[6])
            quotationData.append(['','','','','','MONTO TOTAL',str(finalPrice)])
            exportTable = pd.DataFrame(quotationData,columns=['FECHA','COMPROBANTE','CLIENTE','ESTADO','MONEDA','MONTO DE LA PROFORMA','MONTO (S./)'])
            exportTable.to_excel('CotizacionesInfo.xlsx',index=False)
            doc_excel = openpyxl.load_workbook("CotizacionesInfo.xlsx")
            doc_excel.active.column_dimensions['A'].width = 20
            doc_excel.active.column_dimensions['B'].width = 20
            doc_excel.active.column_dimensions['C'].width = 60
            doc_excel.active.column_dimensions['D'].width = 20
            doc_excel.active.column_dimensions['E'].width = 20
            doc_excel.active.column_dimensions['F'].width = 30
            doc_excel.active.column_dimensions['G'].width = 30
            doc_excel.save("CotizacionesInfo.xlsx")
        else:
            quotationData.append(['INGRESAR AMBAS FECHAS'])
            exportTable = pd.DataFrame(quotationData,columns=['INFORMACION'])
            exportTable.to_excel('CotizacionesInfo.xlsx',index=False)
            doc_excel = openpyxl.load_workbook("CotizacionesInfo.xlsx")
            doc_excel.active.column_dimensions['A'].width = 60
            doc_excel.save("CotizacionesInfo.xlsx")
        response = HttpResponse(open('CotizacionesInfo.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        nombre = 'attachment; ' + 'filename=' + 'CotizacionesInfo.xlsx'
        response['Content-Disposition'] = nombre
        return response

def exportFilteredInvoices(request):
    if request.method == 'POST':
        startDate = request.POST.get('startDate')
        endDate = request.POST.get('endDate')
        invoicesData = []
        if startDate != '' and endDate != '':
            fechaInicio = datetime.datetime.strptime(startDate,'%Y-%m-%d').date()
            fechaFin = datetime.datetime.strptime(endDate,'%Y-%m-%d').date()
            invoicesFilter = invoiceSystem.objects.filter(
                Q(dateInvoice__gte=fechaInicio) &
                Q(dateInvoice__lte=fechaFin)
            ).exclude(stateTeFacturo=None).exclude(stateTeFacturo='Anulada').exclude(stateTeFacturo='').exclude(stateTeFacturo='Rechazado').exclude(stateTeFacturo='Por Anular').order_by('-dateInvoice')
            for invoiceItem in invoicesFilter:
                if len(creditNoteSystem.objects.filter(originCreditNote='INVOICE').filter(asociatedBill=None).exclude(asociatedInvoice=None).filter(asociatedInvoice__codeInvoice=invoiceItem.codeInvoice)) == 0:
                    invoicesData.append([
                        invoiceItem.dateInvoice.strftime('%Y-%m-%d'),
                        invoiceItem.codeInvoice,
                        getInvoiceClientName(invoiceItem),
                        invoiceItem.stateTeFacturo,
                        getInvoiceSellerCode(invoiceItem),
                        getInvoiceListGuides(invoiceItem),
                        invoiceItem.currencyInvoice,
                        getValueInvoice(invoiceItem),
                        getSolesInvoice(invoiceItem),
                    ])
            finalPrice = Decimal(0.00)
            for itemInfo in invoicesData:
                finalPrice = Decimal(finalPrice) + Decimal(itemInfo[8])
            invoicesData.append(['','','','','','','','MONTO TOTAL',str(finalPrice)])
            exportTable = pd.DataFrame(invoicesData,columns=['FECHA','COMPROBANTE','CLIENTE','ESTADO','VENDEDOR','GUIAS','MONEDA','MONTO DE LA FACTURA','MONTO (S./)'])
            exportTable.to_excel('InvoicesInfo.xlsx',index=False)
            doc_excel = openpyxl.load_workbook("InvoicesInfo.xlsx")
            doc_excel.active.column_dimensions['A'].width = 20
            doc_excel.active.column_dimensions['B'].width = 20
            doc_excel.active.column_dimensions['C'].width = 60
            doc_excel.active.column_dimensions['D'].width = 20
            doc_excel.active.column_dimensions['E'].width = 20
            doc_excel.active.column_dimensions['F'].width = 30
            doc_excel.active.column_dimensions['G'].width = 20
            doc_excel.active.column_dimensions['H'].width = 30
            doc_excel.active.column_dimensions['I'].width = 30
            doc_excel.save("InvoicesInfo.xlsx")
        else:
            invoicesData.append(['INGRESAR AMBAS FECHAS'])
            exportTable = pd.DataFrame(invoicesData,columns=['INFORMACION'])
            exportTable.to_excel('InvoicesInfo.xlsx',index=False)
            doc_excel = openpyxl.load_workbook("InvoicesInfo.xlsx")
            doc_excel.active.column_dimensions['A'].width = 60
            doc_excel.save("InvoicesInfo.xlsx")
        response = HttpResponse(open('InvoicesInfo.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        nombre = 'attachment; ' + 'filename=' + 'InvoicesInfo.xlsx'
        response['Content-Disposition'] = nombre
        return response

def getInvoiceSellerCode(invoiceItem):
    sellerCode = ''
    try:
        quotationItem = None
        if invoiceItem.originInvoice == 'GUIDE':
            quotationItem = invoiceItem.guidesystem_set.all()[0].asociatedQuotation
        else:
            quotationItem = invoiceItem.asociatedQuotation
        sellerCode = quotationItem.quotationsellerdata.dataUserQuotation[2]
    except:
        sellerCode = 'NO SELLER'
    return sellerCode

def getInvoiceClientName(invoiceItem):
    clientName = ''
    try:
        quotationItem = None
        if invoiceItem.originInvoice == 'GUIDE':
            quotationItem = invoiceItem.guidesystem_set.all()[0].asociatedQuotation
        else:
            quotationItem = invoiceItem.asociatedQuotation
        clientName = quotationItem.quotationclientdata.dataClientQuotation[1]
    except:
        clientName = 'NO CLIENT'
    return clientName

def getInvoiceListGuides(invoiceItem):
    listGuides = ''
    try:
        totalGuides = invoiceItem.guidesystem_set.all()
        for guideItem in totalGuides:
            listGuides = listGuides + guideItem.codeGuide + ' '
    except:
        listGuides = ''
    return listGuides

def getValueInvoice(invoiceItem):
    valueInvoice = '0.00'
    try:
        quotationItem = None
        if invoiceItem.originInvoice == 'GUIDE':
            quotationItem = invoiceItem.guidesystem_set.all()[0].asociatedQuotation
        else:
            quotationItem = invoiceItem.asociatedQuotation
        valueInvoice = getValueQuotation(quotationItem)
    except:
        valueInvoice = '0.00'
    return valueInvoice


def getSolesInvoice(invoiceItem):
    valueSoles = '0.00'
    try:
        quotationItem = None
        if invoiceItem.originInvoice == 'GUIDE':
            quotationItem = invoiceItem.guidesystem_set.all()[0].asociatedQuotation
        else:
            quotationItem = invoiceItem.asociatedQuotation
        valueSoles = getSolesValue(quotationItem)
    except:
        valueSoles = '0.00'
    return valueSoles



def exportFilteredBills(request):
    if request.method == 'POST':
        startDate = request.POST.get('startDate')
        endDate = request.POST.get('endDate')
        billsData = []
        if startDate != '' and endDate != '':
            fechaInicio = datetime.datetime.strptime(startDate,'%Y-%m-%d').date()
            fechaFin = datetime.datetime.strptime(endDate,'%Y-%m-%d').date()
            billsFilter = billSystem.objects.filter(
                Q(dateBill__gte=fechaInicio) &
                Q(dateBill__lte=fechaFin)
            ).exclude(stateTeFacturo=None).exclude(stateTeFacturo='Anulada').exclude(stateTeFacturo='').exclude(stateTeFacturo='Rechazado').order_by('-dateBill')
            for billItem in billsFilter:
                if len(creditNoteSystem.objects.filter(originCreditNote='BILL').filter(asociatedInvoice=None).exclude(asociatedBill=None).filter(asociatedBill__codeBill=billItem.codeBill)) == 0:
                    billsData.append([
                        billItem.dateBill.strftime('%Y-%m-%d'),
                        billItem.codeBill,
                        getBillClientName(billItem),
                        billItem.stateTeFacturo,
                        getBillSellerCode(billItem),
                        getBillListGuides(billItem),
                        billItem.currencyBill,
                        getValueBill(billItem),
                        getSolesBill(billItem),
                    ])
            finalPrice = Decimal(0.00)
            for itemInfo in billsData:
                finalPrice = Decimal(finalPrice) + Decimal(itemInfo[8])
            billsData.append(['','','','','','','','MONTO TOTAL',str(finalPrice)])
            exportTable = pd.DataFrame(billsData,columns=['FECHA','COMPROBANTE','CLIENTE','ESTADO','VENDEDOR','GUIAS','MONEDA','MONTO DE LA FACTURA','MONTO (S./)'])
            exportTable.to_excel('BillsInfo.xlsx',index=False)
            doc_excel = openpyxl.load_workbook("BillsInfo.xlsx")
            doc_excel.active.column_dimensions['A'].width = 20
            doc_excel.active.column_dimensions['B'].width = 20
            doc_excel.active.column_dimensions['C'].width = 60
            doc_excel.active.column_dimensions['D'].width = 20
            doc_excel.active.column_dimensions['E'].width = 20
            doc_excel.active.column_dimensions['F'].width = 30
            doc_excel.active.column_dimensions['G'].width = 20
            doc_excel.active.column_dimensions['H'].width = 30
            doc_excel.active.column_dimensions['I'].width = 30
            doc_excel.save("BillsInfo.xlsx")
        else:
            billsData.append(['INGRESAR AMBAS FECHAS'])
            exportTable = pd.DataFrame(billsData,columns=['INFORMACION'])
            exportTable.to_excel('BillsInfo.xlsx',index=False)
            doc_excel = openpyxl.load_workbook("BillsInfo.xlsx")
            doc_excel.active.column_dimensions['A'].width = 60
            doc_excel.save("BillsInfo.xlsx")
        response = HttpResponse(open('BillsInfo.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        nombre = 'attachment; ' + 'filename=' + 'BillsInfo.xlsx'
        response['Content-Disposition'] = nombre
        return response


def getBillListGuides(billItem):
    listGuides = ''
    try:
        totalGuides = billItem.guidesystem_set.all()
        for guideItem in totalGuides:
            listGuides = listGuides + guideItem.codeGuide + ' '
    except:
        listGuides = ''
    return listGuides


def getBillSellerCode(billItem):
    sellerCode = ''
    try:
        quotationItem = None
        if billItem.originBill == 'GUIDE':
            quotationItem = billItem.guidesystem_set.all()[0].asociatedQuotation
        else:
            quotationItem = billItem.asociatedQuotation
        sellerCode = quotationItem.quotationsellerdata.dataUserQuotation[2]
    except:
        sellerCode = ''
    return sellerCode

def getBillClientName(billItem):
    clientName = ''
    try:
        quotationItem = None
        if billItem.originBill == 'GUIDE':
            quotationItem = billItem.guidesystem_set.all()[0].asociatedQuotation
        else:
            quotationItem = billItem.asociatedQuotation
        clientName = quotationItem.quotationclientdata.dataClientQuotation[1]
    except:
        clientName = '0.00'
    return clientName

def getValueBill(billItem):
    valueBill = Decimal(0.0000)
    try:
        quotationItem = None
        if billItem.originBill == 'GUIDE':
            for guideItem in billItem.guidesystem_set.all():
                quotationItem = guideItem.asociatedQuotation
                tempValueQuotation = getValueQuotation(quotationItem)
                print(tempValueQuotation)
                valueBill = Decimal(valueBill) + Decimal(tempValueQuotation)
        else:
            quotationItem = billItem.asociatedQuotation
            valueBill = getValueQuotation(quotationItem)
    except:
        valueBill = '0.00'
    valueBill = str(valueBill)
    return valueBill


def getSolesBill(billItem):
    valueSoles = Decimal(0.0000)
    try:
        quotationItem = None
        if billItem.originBill == 'GUIDE':
            for guideItem in billItem.guidesystem_set.all():
                quotationItem = guideItem.asociatedQuotation
                tempValueQuotation = getSolesValue(quotationItem)
                valueSoles = Decimal(valueSoles) + Decimal(tempValueQuotation)
        else:
            quotationItem = billItem.asociatedQuotation
            valueSoles = getSolesValue(quotationItem)
    except:
        valueSoles = '0.00'
    valueSoles = str(valueSoles)
    return valueSoles

def getValueQuotation(quotationItem):
    valueQuotation = Decimal(0.000)
    try:
        if quotationItem.typeQuotation == 'PRODUCTOS':
            print('El error esta en la captura de productos')
            totalProductsQuotation = quotationItem.quotationproductdata_set.all()
            print('Se tienen los productos')
            for productInfo in totalProductsQuotation:
                if quotationItem.currencyQuotation == 'SOLES':
                    if productInfo.dataProductQuotation[5] == 'DOLARES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                    if productInfo.dataProductQuotation[5] == 'SOLES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                if quotationItem.currencyQuotation == 'DOLARES':
                    if productInfo.dataProductQuotation[5] == 'SOLES':
                        v_producto = (Decimal(productInfo.dataProductQuotation[6])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(productInfo.dataProductQuotation[7])/100))
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                    if productInfo.dataProductQuotation[5] == 'DOLARES':
                        v_producto = Decimal(productInfo.dataProductQuotation[6])*Decimal(Decimal(1.00) - Decimal(productInfo.dataProductQuotation[7])/100)
                        v_producto = Decimal('%.2f' % v_producto)*Decimal(productInfo.dataProductQuotation[8])
                if productInfo.dataProductQuotation[9] == '1':
                    v_producto = Decimal(0.00)
                valueQuotation = Decimal(valueQuotation) + Decimal(v_producto)
            print('El error esta en el bucle')
        else:
            totalServicesQuotation = quotationItem.quotationservicedata_set.all()
            for serviceInfo in totalServicesQuotation:
                if quotationItem.currencyQuotation == 'SOLES':
                    if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                        v_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(quotationItem.erSel)*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                    if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                        v_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                if quotationItem.currencyQuotation == 'DOLARES':
                    if serviceInfo.dataServiceQuotation[3] == 'SOLES':
                        v_servicio = (Decimal(serviceInfo.dataServiceQuotation[4])/Decimal(quotationItem.erSel))*Decimal(Decimal(1.00) - (Decimal(serviceInfo.dataServiceQuotation[5])/100))
                    if serviceInfo.dataServiceQuotation[3] == 'DOLARES':
                        v_servicio = Decimal(serviceInfo.dataServiceQuotation[4])*Decimal(Decimal(1.00) - Decimal(serviceInfo.dataServiceQuotation[5])/100)
                valueQuotation = Decimal(valueQuotation) + Decimal(v_servicio)
    except:
        print('Hubo un error')
        valueQuotation = Decimal(0.00)
    valueQuotation = Decimal('%.2f' % valueQuotation)
    valueQuotation = str(valueQuotation)
    return valueQuotation

def getSolesValue(quotationItem):
    valueSoles = Decimal(0.000)
    try:
        if quotationItem.currencyQuotation == 'SOLES':
            valueSoles = Decimal(getValueQuotation(quotationItem))
        else:
            valueSoles = Decimal(round(float(Decimal(getValueQuotation(quotationItem))*Decimal(quotationItem.erSel)),2))
    except:
        valueSoles = Decimal(0.000)
    valueSoles = Decimal('%.2f' % valueSoles)
    valueSoles = str(valueSoles)
    return valueSoles


def exportFilteredGuides(request):
    if request.method == 'POST':
        startDate = request.POST.get('startDate')
        endDate = request.POST.get('endDate')
        guideData = []
        if startDate != '' and endDate != '':
            fechaInicio = datetime.datetime.strptime(startDate,'%Y-%m-%d').date()
            fechaFin = datetime.datetime.strptime(endDate,'%Y-%m-%d').date()
            guideFilter = guideSystem.objects.filter(
                Q(dateGuide__gte=fechaInicio) &
                Q(dateGuide__lte=fechaFin)
            ).order_by('-dateGuide')
            for guideItem in guideFilter:
                guideData.append([
                    guideItem.codeGuide,
                    guideItem.asociatedQuotation.quotationsellerdata.dataUserQuotation[2],
                    guideItem.dateGuide.strftime('%Y-%m-%d'),
                    guideItem.asociatedQuotation.currencyQuotation,
                    guideItem.stateGuide,
                    guideItem.asociatedQuotation.quotationclientdata.dataClientQuotation[1],
                ])
            exportTable = pd.DataFrame(guideData,columns=['CODIGO','VENDEDOR','FECHA','MONEDA','ESTADO','CLIENTE'])
            exportTable.to_excel('GuiasInfo.xlsx',index=False)
            doc_excel = openpyxl.load_workbook("GuiasInfo.xlsx")
            doc_excel.active.column_dimensions['A'].width = 20
            doc_excel.active.column_dimensions['B'].width = 20
            doc_excel.active.column_dimensions['C'].width = 20
            doc_excel.active.column_dimensions['D'].width = 20
            doc_excel.active.column_dimensions['E'].width = 20
            doc_excel.active.column_dimensions['F'].width = 50
            doc_excel.save("GuiasInfo.xlsx")
        else:
            guideData.append(['INGRESAR AMBAS FECHAS'])
            exportTable = pd.DataFrame(guideData,columns=['INFORMACION'])
            exportTable.to_excel('GuiasInfo.xlsx',index=False)
            doc_excel = openpyxl.load_workbook("GuiasInfo.xlsx")
            doc_excel.active.column_dimensions['A'].width = 60
            doc_excel.save("GuiasInfo.xlsx")
        response = HttpResponse(open('GuiasInfo.xlsx','rb'),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        nombre = 'attachment; ' + 'filename=' + 'GuiasInfo.xlsx'
        response['Content-Disposition'] = nombre
        return response

def getBillProducts(request):
    idBill = request.GET.get('idBill')
    billObject = billSystem.objects.get(id=idBill)
    totalProducts = []
    if billObject.originBill == 'GUIDE':
        for guideInfo in billObject.guidesystem_set.all():
            for productInfo in guideInfo.asociatedQuotation.quotationproductdata_set.all():
                totalProducts.append([
                    productInfo.dataProductQuotation[1],
                    productInfo.dataProductQuotation[2],
                    productInfo.dataProductQuotation[8],
                ])
    else:
        for productInfo in billObject.asociatedQuotation.quotationproductdata_set.all():
            totalProducts.append([
                productInfo.dataProductQuotation[1],
                productInfo.dataProductQuotation[2],
                productInfo.dataProductQuotation[8],
            ])
    return JsonResponse({
        'totalProducts':totalProducts
    })

def discountGuideProducts(request,idGuide):
    return HttpResponseRedirect(reverse('salesMetalprotec:guidesMetalprotec'))
